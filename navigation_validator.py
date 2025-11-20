"""
Navigation Menu Validator for Solidigm Website
"""
import time
from typing import Dict, List
from playwright.sync_api import Page
class NavigationValidator:
    def __init__(self, page: Page, base_url: str):
        self.page = page
        self.base_url = base_url
        self.validation_results = []
        self.menu_structure = {
            'main_menus': [],
            'sub_menus': {},
            'broken_links': [],
            'menu_interactions': []
        }
    
    def validate_navigation_menu(self) -> Dict:
        """Validate the complete navigation menu"""
        print("\n" + "="*80)
        print("NAVIGATION MENU VALIDATION")
        print("="*80)
        
        try:
            # Navigate to the page
            print(f"\n[INFO] Navigating to {self.base_url}")
            self.page.goto(self.base_url, timeout=60000, wait_until='domcontentloaded')
            self.page.wait_for_timeout(3000)
            
            # Get page title to verify load
            title = self.page.title()
            print(f"[OK] Page loaded: {title}")
            
            # Start navigation validation
            print("\n[INFO] Starting Navigation Menu Validation...")
            
            # 1. Validate Main Menu Items
            main_menu_items = self._validate_main_menu_items()
            
            # 2. Validate Each Main Menu's Sub-Menus
            sub_menu_results = {}
            for menu_item in main_menu_items:
                print(f"\n[INFO] Checking sub-menus for: {menu_item['name']}")
                sub_results = self._validate_sub_menu(menu_item['name'])
                sub_menu_results[menu_item['name']] = sub_results
            
            # 3. Validate Navigation Links
            link_results = self._validate_all_navigation_links()
            
            # 4. Validate Language Switcher
            language_results = self._validate_language_switcher()
            
            # 5. Validate Search Functionality
            search_results = self._validate_search_functionality()
            
            # 6. Get Navigation Font Styles
            font_styles = self._get_navigation_font_styles()
            
            # Compile results
            results = {
                'url': self.base_url,
                'timestamp': time.strftime("%Y-%m-%d %H:%M:%S"),
                'main_menu_items': main_menu_items,
                'sub_menus': sub_menu_results,
                'link_validation': link_results,
                'language_switcher': language_results,
                'search_functionality': search_results,
                'font_styles': font_styles,
                'summary': self._generate_summary(main_menu_items, sub_menu_results, link_results)
            }
            
            print("\n[SUCCESS] Navigation validation completed!")
            return results
            
        except Exception as e:
            print(f"\n[ERROR] Navigation validation failed: {str(e)}")
            return {
                'url': self.base_url,
                'timestamp': time.strftime("%Y-%m-%d %H:%M:%S"),
                'error': str(e)
            }
    
    def _validate_main_menu_items(self) -> List[Dict]:
        """Validate main menu items presence and visibility"""
        print("\n[MAIN MENU] Validating main menu items...")
        
        # Expected main menu items based on the HTML structure
        expected_menus = ['Product', 'Insights', 'Support', 'Partner', 'Company']
        
        main_menus = []
        
        for menu_name in expected_menus:
            try:
                # Find the menu item
                menu_selector = f'li.cmp-navigation__menu-items'
                all_menu_items = self.page.locator(menu_selector)
                count = all_menu_items.count()
                
                menu_found = False
                menu_item = {}
                
                for i in range(count):
                    item = all_menu_items.nth(i)
                    text = item.locator('.cmp-navigation__menu-text').text_content() or ""
                    
                    if menu_name.lower() in text.lower():
                        is_visible = item.is_visible()
                        link = item.locator('.cmp-navigation__menu-links')
                        has_mega_menu = item.locator('.cmp-navigation__mega-menu').count() > 0
                        
                        menu_item = {
                            'name': menu_name,
                            'text': text.strip(),
                            'is_visible': is_visible,
                            'has_mega_menu': has_mega_menu,
                            'status': 'PASS' if is_visible else 'FAIL'
                        }
                        
                        main_menus.append(menu_item)
                        
                        status_icon = "[OK]" if is_visible else "[FAIL]"
                        print(f"   {status_icon} {menu_name}: Visible={is_visible}, HasMegaMenu={has_mega_menu}")
                        menu_found = True
                        break
                
                if not menu_found:
                    menu_item = {
                        'name': menu_name,
                        'text': '',
                        'is_visible': False,
                        'has_mega_menu': False,
                        'status': 'FAIL'
                    }
                    main_menus.append(menu_item)
                    print(f"   [FAIL] {menu_name}: Not found")
                    
            except Exception as e:
                print(f"   [ERROR] {menu_name}: {str(e)}")
                main_menus.append({
                    'name': menu_name,
                    'error': str(e),
                    'status': 'ERROR'
                })
        
        return main_menus
    
    def _validate_sub_menu(self, menu_name: str) -> List[Dict]:
        """Validate sub-menu items for a given main menu"""
        sub_menu_items = []
        
        try:
            # Hover over the main menu to trigger sub-menu
            menu_selector = f'li.cmp-navigation__menu-items'
            all_menu_items = self.page.locator(menu_selector)
            
            for i in range(all_menu_items.count()):
                item = all_menu_items.nth(i)
                text = item.locator('.cmp-navigation__menu-text').text_content() or ""
                
                if menu_name.lower() in text.lower():
                    # Try to hover to open mega menu
                    item.hover()
                    self.page.wait_for_timeout(1000)
                    
                    # Get all links in the mega menu
                    mega_menu = item.locator('.cmp-navigation__mega-menu')
                    links = mega_menu.locator('.cmp-navigation__mega-menu-links')
                    link_count = links.count()
                    
                    print(f"      Found {link_count} sub-menu items")
                    
                    for j in range(min(link_count, 20)):  # Limit to 20 items
                        link = links.nth(j)
                        link_text = link.text_content() or ""
                        href = link.get_attribute('href') or ""
                        is_visible = link.is_visible() if link_count > 0 else False
                        
                        # Get status code by making a request
                        status_code = 0
                        try:
                            if href and not href.startswith('javascript:'):
                                full_url = self.base_url.rstrip('/') + href if href.startswith('/') else href
                                response = self.page.request.get(full_url, timeout=5000)
                                status_code = response.status
                        except:
                            pass
                        
                        sub_menu_items.append({
                            'text': link_text.strip(),
                            'href': href,
                            'status_code': status_code,
                            'is_visible': is_visible,
                            'status': 'PASS' if status_code == 200 and href else 'FAIL'
                        })
                    
                    break
            
        except Exception as e:
            print(f"      [ERROR] Failed to get sub-menu for {menu_name}: {str(e)}")
        
        return sub_menu_items
    
    def _validate_all_navigation_links(self) -> Dict:
        """Validate all navigation links"""
        print("\n[LINKS] Validating navigation links...")
        
        all_links = []
        broken_links = []
        valid_links = []
        
        try:
            # Get all links in navigation
            nav = self.page.locator('nav')
            links = nav.locator('a[href]')
            count = links.count()
            
            print(f"   Found {count} navigation links")
            
            checked = 0
            for i in range(min(count, 50)):  # Limit to 50 links
                link = links.nth(i)
                href = link.get_attribute('href') or ""
                text = link.text_content() or ""
                is_visible = link.is_visible()
                
                if href:
                    # Skip javascript and # links
                    if href.startswith('javascript:') or href == '#':
                        continue
                    
                    # Validate link
                    try:
                        # Convert relative URLs to absolute
                        from urllib.parse import urljoin
                        absolute_href = href if href.startswith('http') else urljoin(self.page.url, href)
                        
                        response = self.page.request.get(absolute_href, timeout=5000)
                        status_code = response.status
                        is_valid = 200 <= status_code < 400
                        is_broken = status_code >= 400  # Only 4xx and 5xx are truly broken
                        
                        link_info = {
                            'text': text.strip()[:50],
                            'href': href,
                            'status_code': status_code,
                            'is_valid': is_valid,
                            'is_broken': is_broken,
                            'is_visible': is_visible
                        }
                        
                        all_links.append(link_info)
                        
                        if is_valid:
                            valid_links.append(link_info)
                        elif is_broken:  # Only add to broken_links if status_code >= 400
                            broken_links.append(link_info)
                        # Links with status_code 0 (timeout/error) are not added to broken_links
                        
                        checked += 1
                    except Exception as e:
                        # Timeout or other error - don't count as broken, just as not checked
                        link_info = {
                            'text': text.strip()[:50],
                            'href': href,
                            'status_code': 0,
                            'is_valid': False,
                            'is_broken': False,  # Not broken, just not checked
                            'is_visible': is_visible,
                            'error': str(e)[:100]
                        }
                        all_links.append(link_info)
                        # Don't add to broken_links - these are "not checked", not broken
            
            print(f"   Checked: {checked} links")
            print(f"   Valid: {len(valid_links)}")
            print(f"   Broken: {len(broken_links)}")
            
        except Exception as e:
            print(f"   [ERROR] Link validation failed: {str(e)}")
        
        # Separate truly broken links from not checked
        # Note: broken_links already only contains links with status_code >= 400
        truly_broken = broken_links  # No need to filter again, already filtered at line 255-256
        not_checked = [link for link in all_links if link.get('status_code', 0) == 0]
        
        # Debug: Print counts for verification
        print(f"   [DEBUG] Total links: {len(all_links)}, Valid: {len(valid_links)}, Broken: {len(truly_broken)}, Not Checked: {len(not_checked)}")
        
        return {
            'total_checked': len(all_links),
            'valid_links': len(valid_links),
            'broken_links': len(truly_broken),  # Only count actual broken links (4xx, 5xx)
            'broken_details': truly_broken,  # Show ALL broken links, not just first 10
            'not_checked_links': len(not_checked),
            'not_checked_details': not_checked[:20]  # Show first 20 not checked for reference
        }
    
    def _validate_language_switcher(self) -> Dict:
        """Validate language switcher functionality"""
        print("\n[LANGUAGE] Validating language switcher...")
        
        languages = []
        
        try:
            # Find language switcher
            lang_items = self.page.locator('.cmp-navigation__language-items')
            count = lang_items.count()
            
            print(f"   Found {count} language options")
            
            for i in range(count):
                item = lang_items.nth(i)
                link = item.locator('a')
                text = link.text_content() or ""
                href = link.get_attribute('href') or ""
                is_active = 'active' in item.get_attribute('class') or ''
                
                languages.append({
                    'name': text.strip(),
                    'href': href,
                    'is_active': bool(is_active)
                })
                
                status = "[ACTIVE]" if is_active else ""
                print(f"   {status} Language: {text.strip()}")
                
        except Exception as e:
            print(f"   [ERROR] Language switcher validation failed: {str(e)}")
            return {'error': str(e)}
        
        return {
            'total_languages': len(languages),
            'languages': languages
        }
    
    def _validate_search_functionality(self) -> Dict:
        """Validate search functionality"""
        print("\n[SEARCH] Validating search functionality...")
        
        try:
            # Find search button
            search_button = self.page.locator('.c-search__button')
            is_search_visible = search_button.is_visible()
            
            if is_search_visible:
                # Click to open search modal
                search_button.click()
                self.page.wait_for_timeout(1000)
                
                # Check if search modal opened
                search_modal = self.page.locator('.c-search__modal')
                modal_open = search_modal.is_visible()
                
                # Find search input
                search_input = self.page.locator('.c-search-input__field input')
                input_exists = search_input.count() > 0
                
                # Get popular keywords
                popular_keywords = self.page.locator('.modal-tags__links')
                keyword_count = popular_keywords.count()
                
                print(f"   Search visible: {is_search_visible}")
                print(f"   Modal opened: {modal_open}")
                print(f"   Input exists: {input_exists}")
                print(f"   Popular keywords: {keyword_count}")
                
                # Close search modal
                if modal_open:
                    close_button = self.page.locator('.c-search__modal-close')
                    close_button.click()
                    self.page.wait_for_timeout(500)
                
                return {
                    'is_visible': is_search_visible,
                    'modal_opens': modal_open,
                    'input_exists': input_exists,
                    'popular_keywords_count': keyword_count,
                    'status': 'PASS' if is_search_visible and modal_open else 'FAIL'
                }
            else:
                print(f"   [FAIL] Search button not visible")
                return {
                    'is_visible': False,
                    'status': 'FAIL'
                }
                
        except Exception as e:
            print(f"   [ERROR] Search validation failed: {str(e)}")
            return {'error': str(e), 'status': 'ERROR'}
    
    def _get_navigation_font_styles(self) -> Dict:
        """Get font size and color of navigation menu items"""
        print("\n[FONT STYLES] Extracting navigation font styles...")
        
        font_styles = {
            'main_menu': [],
            'sub_menu': [],
            'summary': {}
        }
        
        try:
            # Get main menu styles
            main_menu_items = self.page.locator('li.cmp-navigation__menu-items')
            count = main_menu_items.count()
            
            print(f"   Analyzing {count} main menu items")
            
            for i in range(min(count, 5)):  # Top 5 main menus
                item = main_menu_items.nth(i)
                menu_text = item.locator('.cmp-navigation__menu-text')
                
                # Get computed styles
                styles = self.page.evaluate(f"""
                    () => {{
                        const element = document.querySelectorAll('.cmp-navigation__menu-items')[{i}];
                        if (!element) return null;
                        const textEl = element.querySelector('.cmp-navigation__menu-text');
                        if (!textEl) return null;
                        const styles = window.getComputedStyle(textEl);
                        return {{
                            fontSize: styles.fontSize,
                            fontFamily: styles.fontFamily,
                            fontWeight: styles.fontWeight,
                            color: styles.color,
                            backgroundColor: styles.backgroundColor,
                            textAlign: styles.textAlign
                        }};
                    }}
                """)
                
                if styles:
                    text = menu_text.text_content() or f"Menu {i+1}"
                    font_styles['main_menu'].append({
                        'name': text.strip()[:30],
                        'font_size': styles['fontSize'],
                        'font_color': styles['color'],
                        'font_weight': styles['fontWeight'],
                        'font_family': styles['fontFamily']
                    })
                    
                    print(f"   [OK] {text.strip()[:30]}: {styles['fontSize']}, Color: {styles['color']}")
            
            # Get sub-menu styles (sample)
            mega_menu_links = self.page.locator('.cmp-navigation__mega-menu-links')
            link_count = mega_menu_links.count()
            
            if link_count > 0:
                styles = self.page.evaluate("""
                    () => {
                        const element = document.querySelector('.cmp-navigation__mega-menu-links');
                        if (!element) return null;
                        const styles = window.getComputedStyle(element);
                        return {
                            fontSize: styles.fontSize,
                            fontFamily: styles.fontFamily,
                            fontWeight: styles.fontWeight,
                            color: styles.color,
                            backgroundColor: styles.backgroundColor
                        };
                    }
                """)
                
                if styles:
                    font_styles['sub_menu'] = [{
                        'type': 'sub-menu',
                        'font_size': styles['fontSize'],
                        'font_color': styles['color'],
                        'font_weight': styles['fontWeight'],
                        'font_family': styles['fontFamily']
                    }]
                    print(f"   [OK] Sub-menu links: {styles['fontSize']}, Color: {styles['color']}")
            
            # Generate summary
            if font_styles['main_menu']:
                sample_main = font_styles['main_menu'][0]
                font_styles['summary'] = {
                    'main_menu_font_size': sample_main.get('font_size', ''),
                    'main_menu_font_color': sample_main.get('font_color', ''),
                    'sub_menu_font_size': font_styles['sub_menu'][0].get('font_size', '') if font_styles['sub_menu'] else '',
                    'sub_menu_font_color': font_styles['sub_menu'][0].get('font_color', '') if font_styles['sub_menu'] else ''
                }
        
        except Exception as e:
            print(f"   [ERROR] Failed to get font styles: {str(e)}")
        
        return font_styles
    
    def _generate_summary(self, main_menus, sub_menus, link_results):
        """Generate validation summary"""
        total_main_menus = len(main_menus)
        visible_main_menus = sum(1 for menu in main_menus if menu.get('is_visible', False))
        
        total_sub_items = sum(len(items) for items in sub_menus.values())
        
        return {
            'total_main_menu_items': total_main_menus,
            'visible_main_menu_items': visible_main_menus,
            'total_sub_menu_items': total_sub_items,
            'total_links_checked': link_results.get('total_checked', 0),
            'valid_links': link_results.get('valid_links', 0),
            'broken_links': link_results.get('broken_links', 0)
        }
    
    def generate_excel_report(self, results: Dict) -> str:
        """Generate Excel report for navigation validation"""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"reports/navigation_report_{timestamp}.xlsx"
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary Sheet
        ws = wb.create_sheet("Summary")
        ws['A1'] = "NAVIGATION MENU VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:B1')
        
        summary = results.get('summary', {})
        row = 3
        ws.cell(row=row, column=1, value="Main Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{summary.get('visible_main_menu_items', 0)}/{summary.get('total_main_menu_items', 0)} visible")
        row += 1
        ws.cell(row=row, column=1, value="Sub-Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_sub_menu_items', 0))
        row += 1
        ws.cell(row=row, column=1, value="Links Checked:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_links_checked', 0))
        row += 1
        ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('valid_links', 0))
        row += 1
        ws.cell(row=row, column=1, value="Broken Links:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('broken_links', 0))
        
        # Main Menu Items Sheet
        ws = wb.create_sheet("Main Menu")
        ws['A1'] = "Menu Name"
        ws['B1'] = "Text"
        ws['C1'] = "Is Visible"
        ws['D1'] = "Has Mega Menu"
        ws['E1'] = "Status"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for menu in results.get('main_menu_items', []):
            ws.cell(row=row, column=1, value=menu.get('name', ''))
            ws.cell(row=row, column=2, value=menu.get('text', ''))
            ws.cell(row=row, column=3, value=menu.get('is_visible', False))
            ws.cell(row=row, column=4, value=menu.get('has_mega_menu', False))
            ws.cell(row=row, column=5, value=menu.get('status', ''))
            
            # Color coding
            if menu.get('status') == 'PASS':
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            else:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        # Font Styles Sheet ⭐ NEW!
        font_styles = results.get('font_styles', {})
        if font_styles.get('main_menu') or font_styles.get('sub_menu'):
            ws = wb.create_sheet("Font Styles", 2)
            ws['A1'] = "Element Type"
            ws['B1'] = "Name"
            ws['C1'] = "Font Size"
            ws['D1'] = "Font Color"
            ws['D1'].font = Font(bold=True, color="FFFFFF")
            ws['C1'].font = Font(bold=True, color="FFFFFF")
            
            for cell in ['A1', 'B1', 'C1', 'D1']:
                if ws[cell].font is None or ws[cell].font.color != "FFFFFF":
                    ws[cell].font = Font(bold=True, color="FFFFFF")
                ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row = 2
            
            # Add main menu font styles
            for item in font_styles.get('main_menu', []):
                ws.cell(row=row, column=1, value="Main Menu")
                ws.cell(row=row, column=2, value=item.get('name', ''))
                ws.cell(row=row, column=3, value=item.get('font_size', ''))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                
                row += 1
            
            # Add sub menu font styles
            for item in font_styles.get('sub_menu', []):
                ws.cell(row=row, column=1, value="Sub Menu")
                ws.cell(row=row, column=2, value=item.get('type', ''))
                ws.cell(row=row, column=3, value=item.get('font_size', ''))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                
                row += 1
            
            # Add summary
            summary = font_styles.get('summary', {})
            row += 1
            ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Main Menu Font Size:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary.get('main_menu_font_size', ''))
            row += 1
            ws.cell(row=row, column=1, value="Main Menu Font Color:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary.get('main_menu_font_color', ''))
            row += 1
            ws.cell(row=row, column=1, value="Sub Menu Font Size:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary.get('sub_menu_font_size', ''))
            row += 1
            ws.cell(row=row, column=1, value="Sub Menu Font Color:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary.get('sub_menu_font_color', ''))
            
            # Auto-size columns
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
        
        # Sub-Menu Details Sheet
        ws = wb.create_sheet("Sub-Menu Details")
        ws['A1'] = "Menu"
        ws['B1'] = "Link Text"
        ws['C1'] = "URL"
        ws['D1'] = "Status Code"
        ws['E1'] = "Is Visible"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for menu_name, sub_items in results.get('sub_menus', {}).items():
            for item in sub_items:
                ws.cell(row=row, column=1, value=menu_name)
                ws.cell(row=row, column=2, value=item.get('text', ''))
                ws.cell(row=row, column=3, value=item.get('href', ''))
                
                status_code = item.get('status_code', 0)
                status_cell = ws.cell(row=row, column=4, value=status_code)
                status_cell.number_format = '0'
                
                ws.cell(row=row, column=5, value=item.get('is_visible', False))
                
                # Color coding
                if status_code == 200:
                    fill_color = "D4EDDA"  # Green
                elif status_code == 0:
                    fill_color = "FFF3CD"  # Yellow
                else:
                    fill_color = "F8D7DA"  # Red
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        # Broken Links Sheet
        link_results = results.get('link_validation', {})
        broken_details = link_results.get('broken_details', [])
        if broken_details:
            ws = wb.create_sheet("Broken Links")
            ws['A1'] = "Text"
            ws['B1'] = "URL"
            ws['C1'] = "Status Code"
            ws['D1'] = "Is Visible"
            
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws[cell].font = Font(bold=True, color="FFFFFF")
                ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row = 2
            for link in broken_details:
                ws.cell(row=row, column=1, value=link.get('text', ''))
                ws.cell(row=row, column=2, value=link.get('href', ''))
                status_cell = ws.cell(row=row, column=3, value=link.get('status_code', 0))
                status_cell.number_format = '0'  # Format as number
                ws.cell(row=row, column=4, value=link.get('is_visible', False))
                
                # Color code based on status
                if link.get('status_code', 0) == 200:
                    fill_color = "D4EDDA"  # Green for working links
                elif 400 <= link.get('status_code', 0) < 500:
                    fill_color = "FFE4E1"  # Light red for client errors
                elif link.get('status_code', 0) >= 500:
                    fill_color = "F8D7DA"  # Red for server errors
                else:
                    fill_color = "FFF3CD"  # Yellow for other cases
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        wb.save(filename)
        print(f"\n[EXCEL] Report generated: {filename}")
        return filename

