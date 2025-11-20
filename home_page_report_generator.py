"""
Excel Report Generator for Complete Home Page Validation
"""
import os
from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class HomePageReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def _format_font_size(self, font_size_str: str) -> str:
        """Format font size to 2 decimal places (e.g., '32.0001px' -> '32.00px')"""
        if not font_size_str or not isinstance(font_size_str, str):
            return str(font_size_str) if font_size_str else ''
        
        # Extract numeric value and unit (e.g., "32.0001px" -> "32.00px")
        import re
        match = re.match(r'([\d.]+)(.*)', str(font_size_str).strip())
        if match:
            numeric_value = float(match.group(1))
            unit = match.group(2) if match.group(2) else ''
            return f"{numeric_value:.2f}{unit}"
        return str(font_size_str)
    
    def generate_excel_report(self, results: Dict) -> str:
        """Generate comprehensive Excel report for home page validation"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.output_dir}/homepage_report_{timestamp}.xlsx"
            
            print(f"\n[EXCEL] Generating report: {filename}")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Summary Sheet
            try:
                self._create_summary_sheet(wb, results)
                print(f"[EXCEL] Summary sheet created")
            except Exception as e:
                print(f"[WARNING] Error creating summary sheet: {str(e)}")
            
            # Navigation Sheet
            if results.get('navigation'):
                try:
                    self._create_navigation_sheet(wb, results['navigation'])
                    print(f"[EXCEL] Navigation sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating navigation sheet: {str(e)}")
            
            # Carousel Sheet
            if results.get('carousel'):
                try:
                    self._create_carousel_sheet(wb, results['carousel'])
                    print(f"[EXCEL] Carousel sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating carousel sheet: {str(e)}")
            
            # Featured Products Sheet
            if results.get('featured_products'):
                try:
                    self._create_featured_products_sheet(wb, results['featured_products'])
                    print(f"[EXCEL] Featured Products sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating featured products sheet: {str(e)}")
            
            # Product Cards Sheet - Not present on homepage, skipping
            # if results.get('product_cards'):
            #     try:
            #         self._create_product_cards_sheet(wb, results['product_cards'])
            #         print(f"[EXCEL] Product Cards sheet created")
            #     except Exception as e:
            #         print(f"[WARNING] Error creating product cards sheet: {str(e)}")
            
            # Article List Sheet
            if results.get('article_list'):
                try:
                    self._create_article_list_sheet(wb, results['article_list'])
                    print(f"[EXCEL] Article List sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating article list sheet: {str(e)}")
            
            # Blade Components Sheet
            if results.get('blade_components'):
                try:
                    self._create_blade_components_sheet(wb, results['blade_components'])
                    print(f"[EXCEL] Blade Components sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating blade components sheet: {str(e)}")
            
            # Tile List Sheet
            if results.get('tile_list'):
                try:
                    self._create_tile_list_sheet(wb, results['tile_list'])
                    print(f"[EXCEL] Tile List sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating tile list sheet: {str(e)}")
            
            # Search Sheet
            if results.get('search'):
                try:
                    self._create_search_sheet(wb, results['search'])
                    print(f"[EXCEL] Search sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating search sheet: {str(e)}")
            
            # Footer Sheet (Last tab) - Always create, even if footer data is missing
            try:
                footer_data = results.get('footer', {})
                self._create_footer_sheet(wb, footer_data)
                print(f"[EXCEL] Footer sheet created")
            except Exception as e:
                print(f"[WARNING] Error creating footer sheet: {str(e)}")
            
            # Save the workbook
            try:
                wb.save(filename)
                print(f"\n[EXCEL] [OK] Report successfully saved: {filename}")
                return filename
            except Exception as e:
                print(f"\n[ERROR] Failed to save Excel file: {str(e)}")
                raise
        
        except Exception as e:
            print(f"\n[ERROR] Excel report generation failed: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = "SOLIDIGM HOMEPAGE VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        ws.cell(row=row, column=1, value="URL:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('url', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="Timestamp:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('timestamp', ''))
        row += 2
        
        summary = results.get('summary', {})
        ws.cell(row=row, column=1, value="COMPONENT VALIDATION SUMMARY").font = Font(bold=True, size=12)
        row += 2
        
        # Calculate total slides from all carousels
        carousel_data = results.get('carousel', {})
        total_slides = 0
        carousels = carousel_data.get('carousels', [])
        for carousel in carousels:
            total_slides += carousel.get('slide_count', 0)
        
        # Get featured products count from actual cards
        featured_products_data = results.get('featured_products', {})
        featured_products_count = 0
        if featured_products_data:
            # Check cards.card_count first
            cards_data = featured_products_data.get('cards', {})
            if cards_data:
                featured_products_count = cards_data.get('card_count', 0)
                # If card_count is 0, count actual cards
                if not featured_products_count:
                    actual_cards = cards_data.get('cards', [])
                    featured_products_count = len(actual_cards) if actual_cards else 0
            # Fallback to product_count
            if not featured_products_count:
                featured_products_count = featured_products_data.get('product_count', 0)
        
        components = [
            ("Navigation", summary.get('navigation_validated', False)),
            ("Carousels", f"{total_slides} slides" if total_slides > 0 else "0 found"),
            ("Featured Products", f"{featured_products_count} found"),
            # ("Product Cards", f"{summary.get('product_cards_count', 0)} found"),  # Not on homepage
            ("Articles", f"{summary.get('article_count', 0)} found"),
            ("Blade Components", f"{summary.get('blade_count', 0)} found"),
            ("Tile List", f"{summary.get('tile_list_count', 0)} found"),
            ("Footer", 'Yes' if summary.get('footer_exists') else 'No')
        ]
        
        for component, status in components:
            ws.cell(row=row, column=1, value=component + ":").font = Font(bold=True)
            ws.cell(row=row, column=2, value=status)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _create_navigation_sheet(self, wb: Workbook, nav_results: Dict):
        """Create detailed navigation sheet"""
        ws = wb.create_sheet("Navigation")
        
        # Summary Section
        ws['A1'] = "NAVIGATION MENU SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        summary = nav_results.get('summary', {})
        row = 3
        
        ws.cell(row=row, column=1, value="Total Main Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_main_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Visible Main Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('visible_main_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Sub-Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_sub_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Links Checked:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_links_checked', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
        valid_links = summary.get('valid_links', 0)
        ws.cell(row=row, column=2, value=valid_links)
        if valid_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 1
        
        ws.cell(row=row, column=1, value="Broken Links:").font = Font(bold=True)
        broken_links = summary.get('broken_links', 0)
        ws.cell(row=row, column=2, value=broken_links)
        if broken_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 2
        
        # Main Menu Details Table
        ws.cell(row=row, column=1, value="MAIN MENU DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ["Menu Name", "Display Text", "Is Visible", "Has Mega Menu", "Sub-Menu Count", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        # Get sub-menu counts
        sub_menus = nav_results.get('sub_menus', {})
        main_menu_items = nav_results.get('main_menu_items', [])
        
        for menu in main_menu_items:
            menu_name = menu.get('name', '')
            sub_count = len(sub_menus.get(menu_name, []))
            
            ws.cell(row=row, column=1, value=menu_name)
            ws.cell(row=row, column=2, value=menu.get('text', ''))
            ws.cell(row=row, column=3, value='Yes' if menu.get('is_visible', False) else 'No')
            ws.cell(row=row, column=4, value='Yes' if menu.get('has_mega_menu', False) else 'No')
            ws.cell(row=row, column=5, value=sub_count)
            ws.cell(row=row, column=6, value=menu.get('status', ''))
            
            # Color coding
            if menu.get('status') == 'PASS':
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            else:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        row += 1
        
        # Sub-Menu Details Table
        ws.cell(row=row, column=1, value="SUB-MENU DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ["Main Menu", "Link Text", "URL", "Status Code", "Is Visible", "Link Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for menu_name, sub_items in sub_menus.items():
            for item in sub_items:
                ws.cell(row=row, column=1, value=menu_name)
                ws.cell(row=row, column=2, value=item.get('text', ''))
                ws.cell(row=row, column=3, value=item.get('href', ''))
                
                status_code = item.get('status_code', 0)
                status_cell = ws.cell(row=row, column=4, value=status_code)
                status_cell.number_format = '0'
                
                ws.cell(row=row, column=5, value='Yes' if item.get('is_visible', False) else 'No')
                
                link_status = 'Working' if status_code == 200 else 'Broken' if status_code > 0 else 'Not Checked'
                ws.cell(row=row, column=6, value=link_status)
                
                # Color coding
                if status_code == 200:
                    fill_color = "D4EDDA"  # Green
                elif status_code == 0:
                    fill_color = "FFF3CD"  # Yellow
                else:
                    fill_color = "F8D7DA"  # Red
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        row += 1
        
        # Navigation Links Summary
        link_validation = nav_results.get('link_validation', {})
        if link_validation:
            ws.cell(row=row, column=1, value="NAVIGATION LINKS SUMMARY").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Total Links Checked:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=link_validation.get('total_checked', 0))
            row += 1
            
            ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
            valid = link_validation.get('valid_links', 0)
            ws.cell(row=row, column=2, value=valid)
            if valid > 0:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            row += 1
            
            ws.cell(row=row, column=1, value="Broken Links (4xx/5xx):").font = Font(bold=True)
            broken = link_validation.get('broken_links', 0)
            ws.cell(row=row, column=2, value=broken)
            if broken > 0:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            row += 1
            
            # Not Checked Links (timeouts/errors)
            not_checked = link_validation.get('not_checked_links', 0)
            if not_checked > 0:
                ws.cell(row=row, column=1, value="Not Checked (Timeouts/Errors):").font = Font(bold=True)
                ws.cell(row=row, column=2, value=not_checked)
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                row += 1
            
            row += 1
            
            # Broken Links Details Table (only actual broken links)
            broken_details = link_validation.get('broken_details', [])
            if broken_details:
                ws.cell(row=row, column=1, value="BROKEN LINKS DETAILS").font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
                
                headers = ["Link Text", "URL", "Status Code", "Is Visible", "Error Type"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                row += 1
                
                for link in broken_details:
                    ws.cell(row=row, column=1, value=link.get('text', '')[:50])
                    ws.cell(row=row, column=2, value=link.get('href', '')[:80])
                    status_code = link.get('status_code', 0)
                    ws.cell(row=row, column=3, value=status_code)
                    ws.cell(row=row, column=4, value='Yes' if link.get('is_visible', False) else 'No')
                    ws.cell(row=row, column=5, value='404 Not Found' if status_code == 404 else '403 Forbidden' if status_code == 403 else '500 Server Error' if status_code >= 500 else 'Other Error')
                    
                    # Color coding
                    if status_code == 404:
                        fill_color = "FFE4E1"  # Light red
                    elif status_code == 403:
                        fill_color = "FFB6C1"  # Pink
                    elif status_code >= 500:
                        fill_color = "F8D7DA"  # Red
                    else:
                        fill_color = "FFF3CD"  # Yellow
                    
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    row += 1
            row += 1
        
        # Font Styles Section
        font_styles = nav_results.get('font_styles', {})
        if font_styles.get('main_menu') or font_styles.get('sub_menu'):
            row += 1
            ws.cell(row=row, column=1, value="FONT STYLES").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            headers = ["Element Type", "Name", "Font Size", "Font Color"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            # Main menu font styles
            for item in font_styles.get('main_menu', []):
                ws.cell(row=row, column=1, value="Main Menu")
                ws.cell(row=row, column=2, value=item.get('name', ''))
                ws.cell(row=row, column=3, value=self._format_font_size(item.get('font_size', '')))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                row += 1
            
            # Sub menu font styles
            for item in font_styles.get('sub_menu', []):
                ws.cell(row=row, column=1, value="Sub Menu")
                ws.cell(row=row, column=2, value=item.get('type', ''))
                ws.cell(row=row, column=3, value=self._format_font_size(item.get('font_size', '')))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _create_carousel_sheet(self, wb: Workbook, carousel_results: Dict):
        """Create detailed carousel sheet"""
        ws = wb.create_sheet("Carousels")
        
        # Summary Section
        ws['A1'] = "CAROUSEL SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        row = 3
        carousel_count = len(carousel_results.get('carousels', []))
        ws.cell(row=row, column=1, value="Total Carousels Found:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=carousel_count)
        row += 2
        
        # Carousel Details Table
        ws.cell(row=row, column=1, value="CAROUSEL DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:I{row}')
        row += 1
        
        headers = ["Carousel #", "Title", "Title Font Size", "Title Font Color", "Slides Count", "Container Size (WxH)", "Progress Bar", "Left Chevron", "Right Chevron"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for i, carousel in enumerate(carousel_results.get('carousels', []), 1):
            ws.cell(row=row, column=1, value=f"Carousel {i}")
            
            # Carousel title
            title = carousel.get('title', {})
            ws.cell(row=row, column=2, value=title.get('text', '') if title.get('text') else '')
            ws.cell(row=row, column=3, value=self._format_font_size(title.get('font_size', '')) if title.get('font_size') else '')
            ws.cell(row=row, column=4, value=title.get('font_color', '') if title.get('font_color') else '')
            
            ws.cell(row=row, column=5, value=carousel.get('slide_count', 0))
            
            container = carousel.get('container', {})
            width = float(container.get('width', 0))
            height = float(container.get('height', 0))
            ws.cell(row=row, column=6, value=f"{width:.2f}x{height:.2f}")
            
            pb = carousel.get('progress_bar', {})
            ws.cell(row=row, column=7, value='Yes' if pb.get('exists') else 'No')
            
            nav = carousel.get('navigation', {})
            left_visible = 'Yes' if nav.get('left_chevron_visible') else 'No'
            right_visible = 'Yes' if nav.get('right_chevron_visible') else 'No'
            ws.cell(row=row, column=8, value=left_visible)
            ws.cell(row=row, column=9, value=right_visible)
            
            row += 1
        
        # Slide Details with Text, Font, Buttons, Links
        row += 1
        ws.cell(row=row, column=1, value="SLIDE DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:P{row}')
        row += 1
        
        headers = ["Carousel #", "Slide #", "Title Text", "Title Font Size", "Title Font Color", "Description", "Desc Font Size", "Desc Font Color", "Button Count", "Button Full Text", "Button Link", "Link Valid", "Image URL", "Image Size", "Container Size", "Image Fits"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for i, carousel in enumerate(carousel_results.get('carousels', []), 1):
            slides = carousel.get('slides', [])
            container = carousel.get('container', {})
            width = float(container.get('width', 0))
            height = float(container.get('height', 0))
            container_size = f"{width:.2f}x{height:.2f}"
            
            for slide in slides:
                ws.cell(row=row, column=1, value=f"Carousel {i}")
                ws.cell(row=row, column=2, value=slide.get('index', ''))
                
                # Title and font
                title_text = slide.get('title', '') if isinstance(slide.get('title'), str) else ''
                ws.cell(row=row, column=3, value=title_text if title_text else '')
                
                # Get title font from slide-specific data
                title_font = slide.get('title_font', {})
                title_font_size = self._format_font_size(title_font.get('fontSize', '')) if title_font else ''
                title_font_color = title_font.get('color', '') if title_font else ''
                ws.cell(row=row, column=4, value=title_font_size)
                ws.cell(row=row, column=5, value=title_font_color)
                
                # Description and font
                desc_text = slide.get('description', '') if isinstance(slide.get('description'), str) else ''
                ws.cell(row=row, column=6, value=desc_text if desc_text else '')
                
                # Get description font from slide-specific data
                desc_font = slide.get('description_font', {})
                desc_font_size = self._format_font_size(desc_font.get('fontSize', '')) if desc_font else ''
                desc_font_color = desc_font.get('color', '') if desc_font else ''
                ws.cell(row=row, column=7, value=desc_font_size)
                ws.cell(row=row, column=8, value=desc_font_color)
                
                # Button details
                buttons = slide.get('buttons', [])
                button_count = len(buttons) if buttons else slide.get('button_count', 0)
                ws.cell(row=row, column=9, value=button_count)
                
                # First button details - full text
                if buttons and len(buttons) > 0:
                    first_btn = buttons[0]
                    ws.cell(row=row, column=10, value=first_btn.get('text', ''))  # Full text, no truncation
                    btn_href = first_btn.get('href', '')
                    ws.cell(row=row, column=11, value=btn_href[:80] if btn_href else '')
                    
                    # Link validation status
                    link_valid = 'Valid' if first_btn.get('is_valid') else 'Invalid' if first_btn.get('status_code', 0) > 0 else 'Not Checked'
                    link_valid_cell = ws.cell(row=row, column=12, value=link_valid)
                    if first_btn.get('is_valid'):
                        link_valid_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif first_btn.get('status_code', 0) >= 400:
                        link_valid_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                else:
                    ws.cell(row=row, column=10, value='')
                    ws.cell(row=row, column=11, value='')
                    ws.cell(row=row, column=12, value='')
                
                # Image details
                image_url = slide.get('image_url') or slide.get('main_image') or slide.get('background_image', '')
                ws.cell(row=row, column=13, value=image_url[:80] if image_url else '')
                
                image_size = ''
                if slide.get('image_width') and slide.get('image_height'):
                    img_width = float(slide.get('image_width', 0))
                    img_height = float(slide.get('image_height', 0))
                    image_size = f"{img_width:.2f}x{img_height:.2f}"
                else:
                    image_size = ''
                ws.cell(row=row, column=14, value=image_size)
                
                ws.cell(row=row, column=15, value=container_size)
                
                # Image fits container
                image_fits = slide.get('image_fits_container')
                if image_fits is True:
                    fits_text = 'Yes'
                    fits_color = "D4EDDA"
                elif image_fits is False:
                    fits_text = 'No'
                    fits_color = "F8D7DA"
                else:
                    fits_text = 'Unknown'
                    fits_color = "FFF3CD"
                fits_cell = ws.cell(row=row, column=16, value=fits_text)
                fits_cell.fill = PatternFill(start_color=fits_color, end_color=fits_color, fill_type="solid")
                
                row += 1
        
        row += 1
        
        # Chevron Details
        ws.cell(row=row, column=1, value="CHEVRON DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Carousel #", "Left Chevron", "Right Chevron", "Left Visible", "Right Visible", "Left Clicks", "Right Clicks"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for i, carousel in enumerate(carousel_results.get('carousels', []), 1):
            ws.cell(row=row, column=1, value=f"Carousel {i}")
            
            nav = carousel.get('navigation', {})
            ws.cell(row=row, column=2, value='Yes' if nav.get('has_left_chevron') else 'No')
            ws.cell(row=row, column=3, value='Yes' if nav.get('has_right_chevron') else 'No')
            ws.cell(row=row, column=4, value='Yes' if nav.get('left_chevron_visible') else 'No')
            ws.cell(row=row, column=5, value='Yes' if nav.get('right_chevron_visible') else 'No')
            ws.cell(row=row, column=6, value=f"{nav.get('left_clicks_successful', 0)}/{nav.get('left_clicks_tested', 0)}")
            ws.cell(row=row, column=7, value=f"{nav.get('right_clicks_successful', 0)}/{nav.get('right_clicks_tested', 0)}")
            
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 50
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 60
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 18
        ws.column_dimensions['P'].width = 15
    
    def _create_featured_products_sheet(self, wb: Workbook, fp_results: Dict):
        """Create detailed featured products sheet"""
        ws = wb.create_sheet("Featured Products")
        
        # Summary Section
        ws['A1'] = "FEATURED PRODUCTS SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if fp_results.get('found') or fp_results.get('component_exists') else 'No')
        row += 1
        
        # Get card count from cards data
        cards_data = fp_results.get('cards', {})
        card_count = 0
        if cards_data:
            card_count = cards_data.get('card_count', 0)
            # If card_count is 0, count actual cards
            if not card_count:
                actual_cards = cards_data.get('cards', [])
                card_count = len(actual_cards) if actual_cards else 0
        # Fallback to summary.total_cards
        if not card_count:
            summary = fp_results.get('summary', {})
            card_count = summary.get('total_cards', 0)
        # Last fallback to product_count
        if not card_count:
            card_count = fp_results.get('product_count', 0)
        
        ws.cell(row=row, column=1, value="Total Product Cards:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=card_count)
        row += 2
        
        # Title Details
        title_data = fp_results.get('title', {})
        if title_data:
            ws.cell(row=row, column=1, value="TITLE DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Title Text:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title_data.get('text', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="Font Size:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=self._format_font_size(title_data.get('font_size', '')))
            row += 1
            
            ws.cell(row=row, column=1, value="Font Color:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title_data.get('font_color', ''))
            row += 2
        
        # Product Cards Details
        cards_data = fp_results.get('cards', {})
        products = cards_data.get('cards', []) if cards_data else fp_results.get('products', [])
        if products:
            ws.cell(row=row, column=1, value="PRODUCT CARDS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:H{row}')
            row += 1
            
            headers = ["Card #", "Product Title", "Description", "Image Source", "Image Size", "Container Size", "Link URL", "Link Status", "Title Font", "Desc Font"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for product in products:
                ws.cell(row=row, column=1, value=product.get('index', ''))
                
                title = product.get('title', {}) if isinstance(product.get('title'), dict) else {'text': str(product.get('title', ''))}
                ws.cell(row=row, column=2, value=title.get('text', '')[:50] if title.get('text') else '')
                
                desc = product.get('description', {}) if isinstance(product.get('description'), dict) else {'text': str(product.get('description', ''))}
                ws.cell(row=row, column=3, value=desc.get('text', '')[:50] if desc.get('text') else '')
                
                image = product.get('image', {})
                ws.cell(row=row, column=4, value=image.get('src', '')[:50] if image.get('src') else '')
                if image.get('width') and image.get('height'):
                    img_width = float(image.get('width', 0))
                    img_height = float(image.get('height', 0))
                    ws.cell(row=row, column=5, value=f"{img_width:.2f}x{img_height:.2f}")
                else:
                    ws.cell(row=row, column=5, value='')
                
                container = product.get('container', {})
                width = float(container.get('width', 0))
                height = float(container.get('height', 0))
                ws.cell(row=row, column=6, value=f"{width:.2f}x{height:.2f}")
                
                link = product.get('link', {}) if isinstance(product.get('link'), dict) else {}
                ws.cell(row=row, column=7, value=link.get('href', '')[:50] if link.get('href') else '')
                link_status = 'Valid' if link.get('is_valid') else 'Invalid' if link.get('status_code', 0) > 0 else 'Not Checked'
                ws.cell(row=row, column=8, value=link_status)
                
                # Font styles
                font_styles = product.get('font_styles', {})
                title_font = font_styles.get('title', {})
                title_font_str = f"{self._format_font_size(title_font.get('fontSize', ''))} {title_font.get('color', '')}" if title_font else ''
                ws.cell(row=row, column=9, value=title_font_str[:30])
                
                desc_font = font_styles.get('description', {})
                desc_font_str = f"{self._format_font_size(desc_font.get('fontSize', ''))} {desc_font.get('color', '')}" if desc_font else ''
                ws.cell(row=row, column=10, value=desc_font_str[:30])
                
                row += 1
        
        # Navigation Details
        chevrons = fp_results.get('chevrons', {})
        if chevrons:
            row += 1
            ws.cell(row=row, column=1, value="NAVIGATION DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Left Chevron Works:").font = Font(bold=True)
            ws.cell(row=row, column=2, value='Yes' if chevrons.get('left_works') else 'No')
            row += 1
            
            ws.cell(row=row, column=1, value="Right Chevron Works:").font = Font(bold=True)
            ws.cell(row=row, column=2, value='Yes' if chevrons.get('right_works') else 'No')
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['K'].width = 18
    
    def _create_product_cards_sheet(self, wb: Workbook, pc_results: Dict):
        """Create detailed product cards sheet"""
        ws = wb.create_sheet("Product Cards")
        
        # Summary Section
        ws['A1'] = "PRODUCT CARDS SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if pc_results.get('component_exists') else 'No')
        row += 1
        
        ws.cell(row=row, column=1, value="Total Card Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=pc_results.get('card_count', 0))
        row += 2
        
        # Card Details
        cards = pc_results.get('cards', [])
        if cards:
            ws.cell(row=row, column=1, value="CARD DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            headers = ["Card #", "Card Type", "Container Size", "Link Count", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for i, card in enumerate(cards[:20], 1):  # Limit to first 20 cards
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=card.get('type', 'Product Card') if isinstance(card, dict) else 'Product Card')
                if isinstance(card, dict):
                    container = card.get('container', {})
                    if container:
                        width = float(container.get('width', 0))
                        height = float(container.get('height', 0))
                        ws.cell(row=row, column=3, value=f"{width:.2f}x{height:.2f}")
                    else:
                        ws.cell(row=row, column=3, value='')
                    ws.cell(row=row, column=4, value=card.get('link_count', 0))
                    ws.cell(row=row, column=5, value='Valid' if card.get('is_valid') else 'Invalid')
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_article_list_sheet(self, wb: Workbook, al_results: Dict):
        """Create detailed article list sheet"""
        ws = wb.create_sheet("Article List")
        
        # Summary Section
        ws['A1'] = "ARTICLE LIST SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        cards_data = al_results.get('cards', {})
        card_count = cards_data.get('card_count', 0)
        
        ws.cell(row=row, column=1, value="Component Found:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if al_results.get('found', False) else 'No')
        row += 1
        
        ws.cell(row=row, column=1, value="Total Article Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=card_count)
        row += 1
        
        summary = al_results.get('summary', {})
        ws.cell(row=row, column=1, value="All Links Valid:").font = Font(bold=True)
        links_valid = summary.get('all_links_valid', False)
        ws.cell(row=row, column=2, value='Yes' if links_valid else 'No')
        if links_valid:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        else:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 1
        
        ws.cell(row=row, column=1, value="Chevrons Working:").font = Font(bold=True)
        chevrons_working = summary.get('chevrons_working', False)
        ws.cell(row=row, column=2, value='Yes' if chevrons_working else 'No')
        row += 2
        
        # Title Details
        title_data = al_results.get('title', {})
        if title_data:
            ws.cell(row=row, column=1, value="TITLE DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Title Text:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title_data.get('text', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="View All Link:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title_data.get('view_all_href', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="View All Link Valid:").font = Font(bold=True)
            view_all_valid = title_data.get('view_all_valid', False)
            ws.cell(row=row, column=2, value='Yes' if view_all_valid else 'No')
            row += 2
        
        # Article Cards Details
        cards = cards_data.get('cards', [])
        if cards:
            ws.cell(row=row, column=1, value="ARTICLE CARDS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            headers = ["Card #", "Article Title", "Category", "Image Source", "Image Size", "Container Size", "Link URL", "Link Valid", "Title Font", "Category Font", "Shadow Effect"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for card in cards:
                ws.cell(row=row, column=1, value=card.get('index', ''))
                
                title = card.get('title', {}) if isinstance(card.get('title'), dict) else {'text': str(card.get('title', ''))}
                ws.cell(row=row, column=2, value=title.get('text', '')[:50] if title.get('text') else '')
                
                category = card.get('category', '')
                ws.cell(row=row, column=3, value=category[:30] if category else '')
                
                image = card.get('image', {})
                ws.cell(row=row, column=4, value=image.get('src', '')[:50] if image.get('src') else '')
                if image.get('width') and image.get('height'):
                    img_width = float(image.get('width', 0))
                    img_height = float(image.get('height', 0))
                    ws.cell(row=row, column=5, value=f"{img_width:.2f}x{img_height:.2f}")
                else:
                    ws.cell(row=row, column=5, value='')
                
                container = card.get('container', {})
                width = float(container.get('width', 0))
                height = float(container.get('height', 0))
                ws.cell(row=row, column=6, value=f"{width:.2f}x{height:.2f}")
                
                link = card.get('link', '')
                ws.cell(row=row, column=7, value=link[:50] if link else '')
                
                link_valid = card.get('link_valid', False)
                ws.cell(row=row, column=8, value='Yes' if link_valid else 'No')
                
                # Font styles
                font_styles = card.get('font_styles', {})
                title_font = font_styles.get('title', {})
                title_font_str = f"{title_font.get('fontSize', '')} {title_font.get('color', '')}"[:30] if title_font else ''
                ws.cell(row=row, column=9, value=title_font_str)
                
                category_font = font_styles.get('category', {})
                category_font_str = f"{self._format_font_size(category_font.get('fontSize', ''))} {category_font.get('color', '')}"[:30] if category_font else ''
                ws.cell(row=row, column=10, value=category_font_str)
                
                # Shadow effect - check if card has shadow in hover data
                shadow_effect = 'Detected' if card.get('has_shadow', False) else 'Not Detected'
                ws.cell(row=row, column=11, value=shadow_effect)
                
                row += 1
        
        # Hover/Shadow Effect Details
        hover_data = al_results.get('hover', {})
        if hover_data:
            row += 1
            ws.cell(row=row, column=1, value="SHADOW/HOVER EFFECT DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Hover Effect Detected:").font = Font(bold=True)
            hover_detected = hover_data.get('hover_effect_detected', False)
            ws.cell(row=row, column=2, value='Yes' if hover_detected else 'No')
            if hover_detected:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            row += 1
            
            ws.cell(row=row, column=1, value="Focus Behavior:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=hover_data.get('focus_behavior', '')[:50])
            row += 1
            
            ws.cell(row=row, column=1, value="Is Clickable:").font = Font(bold=True)
            ws.cell(row=row, column=2, value='Yes' if hover_data.get('is_clickable', False) else 'No')
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['K'].width = 18
    
    def _create_blade_components_sheet(self, wb: Workbook, bc_results: Dict):
        """Create detailed blade components sheet"""
        ws = wb.create_sheet("Blade Components")
        
        # Summary Section
        ws['A1'] = "BLADE COMPONENTS SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if bc_results.get('component_exists') else 'No')
        row += 1
        
        ws.cell(row=row, column=1, value="Total Blade Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=bc_results.get('blade_count', 0))
        row += 1
        
        # Image Position Summary
        blades = bc_results.get('blades', [])
        if blades:
            left_count = sum(1 for b in blades if isinstance(b, dict) and b.get('layout') == 'Image Left')
            right_count = sum(1 for b in blades if isinstance(b, dict) and b.get('layout') == 'Image Right')
            ws.cell(row=row, column=1, value="Image Position - Left:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=left_count)
            row += 1
            ws.cell(row=row, column=1, value="Image Position - Right:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=right_count)
        row += 2
        
        # Blade Details
        blades = bc_results.get('blades', [])
        if blades:
            ws.cell(row=row, column=1, value="BLADE DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:M{row}')
            row += 1
            
            headers = ["Blade #", "Image Position", "Container Size", "Image Source", "Image Size", "Title", "Title Font", "Description", "Desc Font", "Button Text", "Button Font", "Button Link", "Link Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for blade in blades:
                if isinstance(blade, dict):
                    ws.cell(row=row, column=1, value=blade.get('index', ''))
                    
                    # Image Position (Layout)
                    image_position = blade.get('layout', 'Unknown')
                    # Make it more explicit
                    if image_position == 'Image Left':
                        image_position = 'Left'
                    elif image_position == 'Image Right':
                        image_position = 'Right'
                    ws.cell(row=row, column=2, value=image_position)
                    
                    # Container size
                    container = blade.get('container', {})
                    if container.get('width') and container.get('height'):
                        container_width = float(container.get('width', 0))
                        container_height = float(container.get('height', 0))
                        ws.cell(row=row, column=3, value=f"{container_width:.2f}x{container_height:.2f}")
                    else:
                        ws.cell(row=row, column=3, value='')
                    
                    # Image details
                    image = blade.get('image', {})
                    ws.cell(row=row, column=4, value=image.get('src', '')[:50] if image.get('src') else '')
                    if image.get('width') and image.get('height'):
                        img_width = float(image.get('width', 0))
                        img_height = float(image.get('height', 0))
                        ws.cell(row=row, column=5, value=f"{img_width:.2f}x{img_height:.2f}")
                    else:
                        ws.cell(row=row, column=5, value='')
                    
                    # Title
                    title = blade.get('title', {})
                    ws.cell(row=row, column=6, value=title.get('text', '')[:40] if title.get('text') else '')
                    title_font = title.get('font_styles', {})
                    title_font_str = f"{title_font.get('fontSize', '')} {title_font.get('color', '')}"[:30] if title_font else ''
                    ws.cell(row=row, column=7, value=title_font_str)
                    
                    # Description
                    description = blade.get('description', {})
                    ws.cell(row=row, column=8, value=description.get('text', '')[:50] if description.get('text') else '')
                    desc_font = description.get('font_styles', {})
                    desc_font_str = f"{desc_font.get('fontSize', '')} {desc_font.get('color', '')}"[:30] if desc_font else ''
                    ws.cell(row=row, column=9, value=desc_font_str)
                    
                    # Button
                    button = blade.get('button', {})
                    ws.cell(row=row, column=10, value=button.get('text', '')[:30] if button.get('text') else '')
                    btn_font = button.get('font_styles', {})
                    btn_font_str = f"{self._format_font_size(btn_font.get('fontSize', ''))} {btn_font.get('color', '')}"[:30] if btn_font else ''
                    ws.cell(row=row, column=11, value=btn_font_str)
                    ws.cell(row=row, column=12, value=button.get('href', '')[:50] if button.get('href') else '')
                    
                    # Link status
                    link_status = 'Valid' if button.get('is_valid') else 'Invalid' if button.get('status_code', 0) > 0 else 'Not Checked'
                    status_cell = ws.cell(row=row, column=13, value=link_status)
                    if button.get('is_valid'):
                        status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif button.get('status_code', 0) > 0:
                        status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    
                    row += 1
            
            row += 1
            
            # Additional Links Table
            ws.cell(row=row, column=1, value="ADDITIONAL LINKS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            headers = ["Blade #", "Link Text", "Link URL", "Status Code", "Link Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for blade in blades:
                if isinstance(blade, dict):
                    blade_index = blade.get('index', '')
                    links = blade.get('links', [])
                    for link in links:
                        if isinstance(link, dict):
                            ws.cell(row=row, column=1, value=blade_index)
                            ws.cell(row=row, column=2, value=link.get('text', '')[:40])
                            ws.cell(row=row, column=3, value=link.get('href', '')[:50])
                            status_code = link.get('status_code', 0)
                            ws.cell(row=row, column=4, value=status_code)
                            link_status = 'Valid' if link.get('is_valid') else 'Invalid' if status_code > 0 else 'Not Checked'
                            ws.cell(row=row, column=5, value=link_status)
                            
                            # Color coding
                            if link.get('is_valid'):
                                fill_color = "D4EDDA"
                            elif status_code > 0:
                                fill_color = "F8D7DA"
                            else:
                                fill_color = "FFF3CD"
                            
                            for col in range(1, 6):
                                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            
                            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 25
        ws.column_dimensions['K'].width = 30
        ws.column_dimensions['L'].width = 50
        ws.column_dimensions['M'].width = 15
    
    def _create_footer_sheet(self, wb: Workbook, footer_results: Dict):
        """Create detailed footer sheet"""
        ws = wb.create_sheet("Footer")
        
        # Summary Section
        ws['A1'] = "FOOTER SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if footer_results.get('component_exists') else 'No')
        row += 1
        
        # Container Size
        container = footer_results.get('container', {})
        if container.get('width') and container.get('height'):
            ws.cell(row=row, column=1, value="Container Size:").font = Font(bold=True)
            container_width = float(container.get('width', 0))
            container_height = float(container.get('height', 0))
            ws.cell(row=row, column=2, value=f"{container_width:.2f}x{container_height:.2f}")
            row += 1
        
        # Number of Columns
        column_count = footer_results.get('column_count', 0) or footer_results.get('section_count', 0)
        ws.cell(row=row, column=1, value="Number of Columns:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=column_count)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Links Count:").font = Font(bold=True)
        total_links = footer_results.get('links_count', 0)
        ws.cell(row=row, column=2, value=total_links)
        row += 1
        
        ws.cell(row=row, column=1, value="Social Icons Count:").font = Font(bold=True)
        social_count = footer_results.get('social_icon_count', 0)
        ws.cell(row=row, column=2, value=social_count)
        row += 1
        
        # Calculate valid and broken links
        valid_links = 0
        broken_links = 0
        not_checked_links = 0
        
        # Count from section links
        sections = footer_results.get('sections', [])
        for section in sections:
            links = section.get('links', [])
            for link in links:
                status_code = link.get('status_code', 0)
                if link.get('is_valid'):
                    valid_links += 1
                elif status_code >= 400:
                    broken_links += 1
                elif status_code == 0:
                    not_checked_links += 1
        
        # Count from social icons
        social_icons = footer_results.get('social_icons', [])
        for icon in social_icons:
            status_code = icon.get('status_code', 0)
            if icon.get('is_valid'):
                valid_links += 1
            elif status_code >= 400:
                broken_links += 1
            elif status_code == 0:
                not_checked_links += 1
        
        ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=valid_links)
        if valid_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 1
        
        ws.cell(row=row, column=1, value="Broken Links (4xx/5xx):").font = Font(bold=True)
        ws.cell(row=row, column=2, value=broken_links)
        if broken_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 1
        
        if not_checked_links > 0:
            ws.cell(row=row, column=1, value="Not Checked (Timeouts/Errors):").font = Font(bold=True)
            ws.cell(row=row, column=2, value=not_checked_links)
            ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            row += 1
        
        row += 1
        
        # Logo Details
        logo = footer_results.get('logo', {})
        if logo.get('href'):
            ws.cell(row=row, column=1, value="LOGO DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Logo Link:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=logo.get('href', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="Logo Title:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=logo.get('title', ''))
            row += 2
        
        # Social Icons Details
        social_icons = footer_results.get('social_icons', [])
        if social_icons:
            ws.cell(row=row, column=1, value="SOCIAL ICONS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:I{row}')
            row += 1
            
            headers = ["Icon #", "Aria Label", "Link URL", "Target", "Is Clickable", "Domain Valid", "Domain Validation", "Status Code", "Link Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for icon in social_icons:
                ws.cell(row=row, column=1, value=icon.get('index', ''))
                ws.cell(row=row, column=2, value=icon.get('aria_label', '')[:40])
                ws.cell(row=row, column=3, value=icon.get('href', '')[:60])
                ws.cell(row=row, column=4, value=icon.get('target', '_self'))
                
                # Check if icon link is clickable
                is_clickable = False
                try:
                    # Social icons should be clickable if they have a valid href
                    is_clickable = bool(icon.get('href')) and icon.get('href') != '#' and not icon.get('href', '').startswith('#')
                except:
                    pass
                
                clickable_cell = ws.cell(row=row, column=5, value='Yes' if is_clickable else 'No')
                if is_clickable:
                    clickable_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                else:
                    clickable_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                
                # Domain validation
                domain_valid = icon.get('domain_valid', False)
                domain_cell = ws.cell(row=row, column=6, value='Yes' if domain_valid else 'No')
                if domain_valid:
                    domain_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                else:
                    domain_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                
                ws.cell(row=row, column=7, value=icon.get('domain_validation_message', '')[:50])
                
                status_code = icon.get('status_code', 0)
                ws.cell(row=row, column=8, value=status_code)
                
                link_status = 'Valid' if icon.get('is_valid') else 'Invalid' if status_code > 0 else 'Not Checked'
                status_cell = ws.cell(row=row, column=9, value=link_status)
                
                # Color coding for link status
                if icon.get('is_valid'):
                    status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif status_code >= 400:
                    status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                
                row += 1
            row += 1
        
        # Copyright Details
        copyright = footer_results.get('copyright', {})
        if copyright.get('text'):
            ws.cell(row=row, column=1, value="COPYRIGHT DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Copyright Text:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=copyright.get('text', '')[:100])
            row += 2
        
        # Left Column Details (Column 1)
        logo = footer_results.get('logo', {})
        social_icons = footer_results.get('social_icons', [])
        copyright = footer_results.get('copyright', {})
        
        if logo.get('href') or social_icons or copyright.get('text'):
            ws.cell(row=row, column=1, value="COLUMN 1 DETAILS (Left Column)").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Column #:").font = Font(bold=True)
            ws.cell(row=row, column=2, value="1")
            row += 1
            
            ws.cell(row=row, column=1, value="Column Heading:").font = Font(bold=True)
            ws.cell(row=row, column=2, value="Logo, Social Icons, Copyright")
            row += 1
            
            ws.cell(row=row, column=1, value="Sub-Menu Count:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(social_icons))
            row += 2
        
        # Footer Sections Details (Columns 2, 3, 4, etc.)
        sections = footer_results.get('sections', [])
        if sections:
            ws.cell(row=row, column=1, value="NAVIGATION SECTIONS DETAILS (Columns 2+)").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            headers = ["Column #", "Column Heading", "Sub-Menu Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for section in sections:
                if isinstance(section, dict):
                    ws.cell(row=row, column=1, value=section.get('index', ''))
                    section_title = section.get('title', '')
                    # Handle empty titles - show "(No Title)" if empty
                    if not section_title or section_title == '(No Title)':
                        display_title = '(No Title)'
                    else:
                        display_title = section_title[:40]
                    ws.cell(row=row, column=2, value=display_title)
                    ws.cell(row=row, column=3, value=section.get('link_count', 0))
                    row += 1
            
            row += 1
            
            # Section Links Details
            ws.cell(row=row, column=1, value="SECTION LINKS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:I{row}')
            row += 1
            
            headers = ["Column #", "Column Heading", "Sub-Menu Count", "Link Text", "Link URL", "Status Code", "Link Status", "Font Size", "Font Color"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for section in sections:
                if isinstance(section, dict):
                    section_index = section.get('index', '')
                    section_title = section.get('title', '')
                    links = section.get('links', [])
                    link_count = len(links)  # Sub-menu count for this section
                    
                    for link in links:
                        if isinstance(link, dict):
                            ws.cell(row=row, column=1, value=section_index)
                            # Handle empty titles
                            if not section_title or section_title == '(No Title)':
                                display_title = '(No Title)'
                            else:
                                display_title = section_title[:30]
                            ws.cell(row=row, column=2, value=display_title)
                            ws.cell(row=row, column=3, value=link_count)  # Sub-menu count
                            ws.cell(row=row, column=4, value=link.get('text', '')[:40])
                            ws.cell(row=row, column=5, value=link.get('href', '')[:60])
                            
                            status_code = link.get('status_code', 0)
                            ws.cell(row=row, column=6, value=status_code)
                            
                            link_status = 'Valid' if link.get('is_valid') else 'Invalid' if status_code > 0 else 'Not Checked'
                            status_cell = ws.cell(row=row, column=7, value=link_status)
                            
                            # Color coding for link status
                            if link.get('is_valid'):
                                status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            elif status_code >= 400:
                                status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                            else:
                                status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            
                            # Font styles
                            font_styles = link.get('font_styles', {})
                            if font_styles:
                                font_size = self._format_font_size(font_styles.get('fontSize', ''))
                                font_color = font_styles.get('color', '')
                            else:
                                # Try to get from direct properties
                                font_size = self._format_font_size(link.get('font_size', ''))
                                font_color = link.get('font_color', '')
                            
                            ws.cell(row=row, column=8, value=font_size)
                            ws.cell(row=row, column=9, value=font_color)
                            
                            row += 1
        
        # Footer Links Summary (similar to Navigation Links Summary)
        row += 1
        ws.cell(row=row, column=1, value="FOOTER LINKS SUMMARY").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        total_footer_links = total_links + social_count
        ws.cell(row=row, column=1, value="Total Links Checked:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_footer_links)
        row += 1
        
        ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=valid_links)
        if valid_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 1
        
        ws.cell(row=row, column=1, value="Broken Links (4xx/5xx):").font = Font(bold=True)
        ws.cell(row=row, column=2, value=broken_links)
        if broken_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 1
        
        if not_checked_links > 0:
            ws.cell(row=row, column=1, value="Not Checked (Timeouts/Errors):").font = Font(bold=True)
            ws.cell(row=row, column=2, value=not_checked_links)
            ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            row += 1
        
        row += 1
        
        # Broken Links Details Table (similar to Navigation Broken Links)
        broken_links_list = []
        
        # Collect broken links from sections
        for section in sections:
            links = section.get('links', [])
            for link in links:
                status_code = link.get('status_code', 0)
                if status_code >= 400:
                    broken_links_list.append({
                        'text': link.get('text', ''),
                        'href': link.get('href', ''),
                        'status_code': status_code,
                        'section': section.get('title', ''),
                        'is_clickable': link.get('is_clickable', False)
                    })
        
        # Collect broken links from social icons
        for icon in social_icons:
            status_code = icon.get('status_code', 0)
            if status_code >= 400:
                broken_links_list.append({
                    'text': icon.get('aria_label', 'Social Icon'),
                    'href': icon.get('href', ''),
                    'status_code': status_code,
                    'section': 'Social Icons',
                    'is_clickable': bool(icon.get('href'))
                })
        
        if broken_links_list:
            ws.cell(row=row, column=1, value="BROKEN LINKS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            headers = ["Link Text", "URL", "Status Code", "Section", "Is Clickable", "Error Type"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for link in broken_links_list:
                ws.cell(row=row, column=1, value=link.get('text', '')[:50])
                ws.cell(row=row, column=2, value=link.get('href', '')[:80])
                status_code = link.get('status_code', 0)
                ws.cell(row=row, column=3, value=status_code)
                ws.cell(row=row, column=4, value=link.get('section', '')[:30])
                ws.cell(row=row, column=5, value='Yes' if link.get('is_clickable') else 'No')
                
                # Error type
                if status_code == 404:
                    error_type = '404 Not Found'
                    fill_color = "FFE4E1"
                elif status_code == 403:
                    error_type = '403 Forbidden'
                    fill_color = "FFB6C1"
                elif status_code >= 500:
                    error_type = '500 Server Error'
                    fill_color = "F8D7DA"
                else:
                    error_type = 'Other Error'
                    fill_color = "FFF3CD"
                
                ws.cell(row=row, column=6, value=error_type)
                
                # Color coding
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        row += 1
        
        # Trademark Details
        trademark = footer_results.get('trademark', {})
        if trademark.get('text') or trademark.get('trustarc_link'):
            row += 1
            ws.cell(row=row, column=1, value="TRADEMARK DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            if trademark.get('text'):
                ws.cell(row=row, column=1, value="Trademark Text:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=trademark.get('text', '')[:100])
                row += 1
            
            if trademark.get('trustarc_link'):
                ws.cell(row=row, column=1, value="TrustArc Link:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=trademark.get('trustarc_link', ''))
                row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
    
    def _create_tile_list_sheet(self, wb: Workbook, tl_results: Dict):
        """Create detailed tile list sheet"""
        ws = wb.create_sheet("Tile List")
        
        # Summary Section
        ws['A1'] = "TILE LIST SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if tl_results.get('component_exists') else 'No')
        row += 1
        
        ws.cell(row=row, column=1, value="Total Tiles Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=tl_results.get('tile_count', 0))
        row += 2
        
        # Big Container Details
        big_container = tl_results.get('big_container', {})
        if big_container:
            ws.cell(row=row, column=1, value="BIG CONTAINER DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Container ID:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=big_container.get('id', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="Container Size:").font = Font(bold=True)
            if big_container.get('width') and big_container.get('height'):
                container_width = float(big_container.get('width', 0))
                container_height = float(big_container.get('height', 0))
                ws.cell(row=row, column=2, value=f"{container_width:.2f}x{container_height:.2f}")
            else:
                ws.cell(row=row, column=2, value='')
            row += 1
            
            ws.cell(row=row, column=1, value="Container Background Color:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=big_container.get('background_color', ''))
            row += 2
        
        # Tile Details Table
        tiles = tl_results.get('tiles', [])
        if tiles:
            ws.cell(row=row, column=1, value="TILE DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:M{row}')
            row += 1
            
            headers = ["Tile #", "Title Text", "Text Size", "Text Color", "Tile Container Size", "Tile Container Color", "Icon URL", "Icon Size", "Link URL", "Is Clickable", "Link Status", "Target"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for tile in tiles:
                if isinstance(tile, dict):
                    ws.cell(row=row, column=1, value=tile.get('index', ''))
                    
                    # Title (Text)
                    title = tile.get('title', {})
                    ws.cell(row=row, column=2, value=title.get('text', '')[:40] if title.get('text') else '')
                    ws.cell(row=row, column=3, value=self._format_font_size(title.get('font_size', '')))
                    ws.cell(row=row, column=4, value=title.get('font_color', ''))
                    
                    # Tile Container
                    container = tile.get('container', {})
                    if container.get('width') and container.get('height'):
                        tile_width = float(container.get('width', 0))
                        tile_height = float(container.get('height', 0))
                        ws.cell(row=row, column=5, value=f"{tile_width:.2f}x{tile_height:.2f}")
                    else:
                        ws.cell(row=row, column=5, value='')
                    ws.cell(row=row, column=6, value=container.get('background_color', ''))
                    
                    # Icon
                    icon = tile.get('icon', {})
                    ws.cell(row=row, column=7, value=icon.get('url', '')[:60] if icon.get('url') else '')
                    if icon.get('width') and icon.get('height'):
                        icon_width = float(icon.get('width', 0))
                        icon_height = float(icon.get('height', 0))
                        ws.cell(row=row, column=8, value=f"{icon_width:.2f}x{icon_height:.2f}")
                    else:
                        ws.cell(row=row, column=8, value='')
                    
                    # Link
                    link = tile.get('link', {})
                    ws.cell(row=row, column=9, value=link.get('href', '')[:60] if link.get('href') else '')
                    
                    # Is Clickable
                    is_clickable = link.get('is_clickable', False)
                    clickable_cell = ws.cell(row=row, column=10, value='Yes' if is_clickable else 'No')
                    if is_clickable:
                        clickable_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    else:
                        clickable_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    
                    # Link status (should not be broken)
                    link_status = 'Valid' if link.get('is_valid') else 'Invalid' if link.get('status_code', 0) > 0 else 'Not Checked'
                    status_cell = ws.cell(row=row, column=11, value=link_status)
                    if link.get('is_valid'):
                        status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif link.get('status_code', 0) >= 400:
                        status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    else:
                        status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    
                    ws.cell(row=row, column=12, value=link.get('target', '_self'))
                    
                    row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 60
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 60
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 12
    
    def _create_search_sheet(self, wb: Workbook, search_results: Dict):
        """Create detailed search sheet"""
        ws = wb.create_sheet("Search")
        
        # Summary Section
        ws['A1'] = "SEARCH COMPONENT SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:C1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Exists:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if search_results.get('component_exists') else 'No')
        row += 1
        
        suggestion_count = search_results.get('suggestion_count', 0)
        ws.cell(row=row, column=1, value="Search Suggestions Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=suggestion_count)
        row += 2
        
        # Title Details
        title = search_results.get('title', {})
        if title.get('text'):
            ws.cell(row=row, column=1, value="TITLE DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Title Text:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title.get('text', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="Font Size:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=self._format_font_size(title.get('font_size', '')))
            row += 1
            
            ws.cell(row=row, column=1, value="Font Color:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=title.get('font_color', ''))
            row += 2
        
        # Form Details
        form = search_results.get('form', {})
        if form.get('action'):
            ws.cell(row=row, column=1, value="FORM DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Form Action:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=form.get('action', ''))
            row += 1
            
            ws.cell(row=row, column=1, value="Form Method:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=form.get('method', 'get'))
            row += 1
            
            # Input field details
            input_field = form.get('input', {})
            if input_field:
                ws.cell(row=row, column=1, value="Input Placeholder:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=input_field.get('placeholder', ''))
                row += 1
                
                ws.cell(row=row, column=1, value="Is Required:").font = Font(bold=True)
                ws.cell(row=row, column=2, value='Yes' if input_field.get('required') else 'No')
                row += 1
                
                if input_field.get('data_search_page'):
                    ws.cell(row=row, column=1, value="Data Search Page:").font = Font(bold=True)
                    ws.cell(row=row, column=2, value=input_field.get('data_search_page', ''))
                    row += 1
            
            row += 1
        
        # Search Suggestions Details
        suggestions = search_results.get('suggestions', [])
        if suggestions:
            ws.cell(row=row, column=1, value="SEARCH SUGGESTIONS DETAILS").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            headers = ["Suggestion #", "Suggestion Text", "Link URL", "Status Code", "Link Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            for suggestion in suggestions:
                ws.cell(row=row, column=1, value=suggestion.get('index', ''))
                ws.cell(row=row, column=2, value=suggestion.get('text', '')[:40])
                ws.cell(row=row, column=3, value=suggestion.get('href', '')[:60])
                
                status_code = suggestion.get('status_code', 0)
                ws.cell(row=row, column=4, value=status_code)
                
                link_status = 'Valid' if suggestion.get('is_valid') else 'Invalid' if status_code > 0 else 'Not Checked'
                status_cell = ws.cell(row=row, column=5, value=link_status)
                
                # Color coding
                if suggestion.get('is_valid'):
                    status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif status_code >= 400:
                    status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15


