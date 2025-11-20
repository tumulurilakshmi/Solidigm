"""
Excel Report Generator for Data Center Landing Page Validation
"""
import os
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DataCenterPageReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def _format_font_size(self, font_size_str: str) -> str:
        """Format font size to 2 decimal places (e.g., '32.0001px' -> '32.00px')"""
        if not font_size_str or not isinstance(font_size_str, str):
            return str(font_size_str) if font_size_str else ''
        
        # Extract numeric value and unit (e.g., "32.0001px" -> "32.00px")
        import re
        match = re.match(r'([\d.]+)(.*)', str(font_size_str).strip())
        if match:
            numeric_value = float(match.group(1))
            unit = match.group(2) if match.group(2) else ''
            return f"{numeric_value:.2f}{unit}"
        return str(font_size_str)
    
    def generate_excel_report(self, results: Dict) -> str:
        """Generate Excel report for data center page validation"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.output_dir}/data_center_page_validation_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Center Page"
        
        # Create summary sheet
        self._create_summary_sheet(ws, results)
        
        wb.save(filename)
        print(f"\n[SUCCESS] Excel report saved: {filename}")
        return filename
    
    def _create_summary_sheet(self, ws, results: Dict):
        """Create summary sheet with all validation details"""
        row = 1
        
        # Title
        ws.cell(row, 1, "Data Center Landing Page Validation Report").font = Font(bold=True, size=16)
        ws.merge_cells(f'A{row}:D{row}')
        row += 2
        
        # URL
        ws.cell(row, 1, "URL:").font = Font(bold=True)
        ws.cell(row, 2, results.get('url', ''))
        row += 1
        
        # Validation Date
        ws.cell(row, 1, "Validation Date:").font = Font(bold=True)
        ws.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 2
        
        # Summary Section
        ws.cell(row, 1, "Summary").font = Font(bold=True, size=14)
        row += 1
        
        summary = results.get('summary', {})
        summary_items = [
            ("Header Found", summary.get('header_found', False)),
            ("Footer Found", summary.get('footer_found', False)),
            ("Hero Component Found", summary.get('hero_found', False)),
            ("Series Cards Found", summary.get('series_cards_found', False)),
            ("All Series Present (D7, D5, D3)", summary.get('all_series_present', False)),
            ("Model List Found", summary.get('model_list_found', False)),
            ("Related Articles Found", summary.get('articles_found', False))
        ]
        
        for label, value in summary_items:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, 'Yes' if value else 'No')
            ws.cell(row, 2).fill = PatternFill(start_color="C6EFCE" if value else "FFC7CE", 
                                               end_color="C6EFCE" if value else "FFC7CE", 
                                               fill_type="solid")
            row += 1
        
        row += 2
        
        # Header/Footer Section
        header_footer = results.get('header_footer', {})
        if header_footer:
            ws.cell(row, 1, "Header & Footer").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row, 1, "Header Found:")
            ws.cell(row, 2, 'Yes' if header_footer.get('header_found') else 'No')
            row += 1
            ws.cell(row, 1, "Footer Found:")
            ws.cell(row, 2, 'Yes' if header_footer.get('footer_found') else 'No')
            row += 2
        
        # Hero Component Section
        hero = results.get('hero', {})
        if hero.get('found'):
            ws.cell(row, 1, "Hero Component").font = Font(bold=True, size=12)
            row += 1
            
            container = hero.get('container', {})
            if container.get('found'):
                ws.cell(row, 1, "Container Size:")
                width = float(container.get('width', 0))
                height = float(container.get('height', 0))
                ws.cell(row, 2, f"{width:.2f}x{height:.2f} px")
                row += 1
            
            title = hero.get('title', {})
            if title.get('found'):
                ws.cell(row, 1, "Hero Title:")
                ws.cell(row, 2, title.get('text', ''))
                row += 1
                ws.cell(row, 1, "Title Font Size:")
                ws.cell(row, 2, self._format_font_size(title.get('font_size', '')))
                row += 1
                ws.cell(row, 1, "Title Font Color:")
                ws.cell(row, 2, title.get('font_color', ''))
                row += 1
            
            description = hero.get('description', {})
            if description.get('found'):
                ws.cell(row, 1, "Description:")
                ws.cell(row, 2, description.get('text', '')[:200])
                row += 1
            
            row += 1
        
        # Series Cards Section
        series_cards = results.get('series_cards', {})
        if series_cards.get('found'):
            ws.cell(row, 1, "Series Cards").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row, 1, "All Series Present:")
            ws.cell(row, 2, 'Yes' if series_cards.get('all_series_present') else 'No')
            row += 1
            ws.cell(row, 1, "Series Cards Found:")
            ws.cell(row, 2, len(series_cards.get('cards', [])))
            row += 2
            
            # Series cards details table
            if series_cards.get('cards'):
                headers = ['Series', 'Title', 'Href', 'URL Format Valid', 'Notes']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                row += 1
                
                for card in series_cards.get('cards', []):
                    ws.cell(row, 1, card.get('series', ''))
                    ws.cell(row, 2, card.get('title', ''))
                    ws.cell(row, 3, card.get('href', ''))
                    url_valid = card.get('url_format_valid', False)
                    ws.cell(row, 4, 'Yes' if url_valid else 'No')
                    ws.cell(row, 4).fill = PatternFill(start_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       end_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       fill_type="solid")
                    ws.cell(row, 5, 'N/A')  # Navigation no longer tested
                    row += 1
                row += 1
        
        # Model List Section
        model_list = results.get('model_list', {})
        if model_list.get('found'):
            ws.cell(row, 1, "Model List Section").font = Font(bold=True, size=12)
            row += 1
            
            title = model_list.get('title', {})
            if title.get('found'):
                ws.cell(row, 1, "Title:")
                ws.cell(row, 2, title.get('text', ''))
                row += 1
            
            default_cards = model_list.get('default_cards', {})
            ws.cell(row, 1, "Default Product Cards Count:")
            ws.cell(row, 2, default_cards.get('card_count', 0))
            row += 2
            
            # Default Product Cards Details Table
            default_cards_list = default_cards.get('cards', [])
            if default_cards_list:
                ws.cell(row, 1, "Default Product Cards Details").font = Font(bold=True, size=11)
                row += 1
                
                headers = ['Index', 'Product Title', 'Title Font', 'Title Color', 'Description', 'Description Font', 
                          'Container Size', 'Image Src', 'Interface', 'Interface Font', 'Form Factor', 'Form Factor Font',
                          'Capacity', 'Capacity Font', 'View Details Link', 'View Details Button Text', 'Button Font', 
                          'Button BG Color', 'Compare Button Text', 'URL Format Valid']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row += 1
                
                for card in default_cards_list:
                    ws.cell(row, 1, card.get('index', ''))
                    
                    # Title
                    title_text = card.get('title', {}).get('text', '')
                    ws.cell(row, 2, title_text)
                    ws.cell(row, 3, card.get('title', {}).get('fontSize', ''))
                    ws.cell(row, 4, card.get('title', {}).get('color', ''))
                    
                    # Description
                    desc_text = card.get('description', {}).get('text', '')
                    ws.cell(row, 5, desc_text[:100] + '...' if len(desc_text) > 100 else desc_text)
                    ws.cell(row, 6, self._format_font_size(card.get('description', {}).get('fontSize', '')))
                    
                    # Container
                    container = card.get('container', {})
                    width = float(container.get('width', 0))
                    height = float(container.get('height', 0))
                    container_size = f"{width:.2f}x{height:.2f}"
                    ws.cell(row, 7, container_size)
                    
                    # Image
                    image_src = card.get('image', {}).get('src', '')
                    ws.cell(row, 8, image_src[:50] + '...' if len(image_src) > 50 else image_src)
                    
                    # Interface
                    interface_text = card.get('interface', {}).get('text', '')
                    ws.cell(row, 9, interface_text)
                    ws.cell(row, 10, self._format_font_size(card.get('interface', {}).get('fontSize', '')))
                    
                    # Form Factor
                    form_factor_text = card.get('form_factor', {}).get('text', '')
                    ws.cell(row, 11, form_factor_text)
                    ws.cell(row, 12, self._format_font_size(card.get('form_factor', {}).get('fontSize', '')))
                    
                    # Capacity
                    capacity_text = card.get('capacity', {}).get('text', '')
                    ws.cell(row, 13, capacity_text)
                    ws.cell(row, 14, self._format_font_size(card.get('capacity', {}).get('fontSize', '')))
                    
                    # View Details
                    view_details_link = card.get('view_details_link', '')
                    ws.cell(row, 15, view_details_link)
                    view_details_btn = card.get('view_details_button', {})
                    ws.cell(row, 16, view_details_btn.get('text', ''))
                    ws.cell(row, 17, self._format_font_size(view_details_btn.get('fontSize', '')))
                    ws.cell(row, 18, view_details_btn.get('backgroundColor', ''))
                    
                    # Compare Button
                    compare_btn = card.get('compare_button', {})
                    ws.cell(row, 19, compare_btn.get('text', ''))
                    
                    # URL Format Valid
                    url_valid = card.get('url_format_valid', False)
                    ws.cell(row, 20, 'Yes' if url_valid else 'No')
                    ws.cell(row, 20).fill = PatternFill(start_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       end_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       fill_type="solid")
                    row += 1
                
                row += 1
            
            # Filtering Section
            filtered_cards = model_list.get('filtered_cards', {})
            if filtered_cards:
                ws.cell(row, 1, "Filtering Test Results").font = Font(bold=True, size=11)
                row += 1
                
                # Show selected filters
                selected_filters = filtered_cards.get('selected_filters', {})
                if selected_filters:
                    ws.cell(row, 1, "Selected Filters:").font = Font(bold=True)
                    row += 1
                    
                    if selected_filters.get('interface'):
                        ws.cell(row, 1, "  Interface:")
                        ws.cell(row, 2, selected_filters.get('interface'))
                        row += 1
                    
                    if selected_filters.get('form_factor'):
                        ws.cell(row, 1, "  Form Factor:")
                        ws.cell(row, 2, selected_filters.get('form_factor'))
                        row += 1
                    
                    if selected_filters.get('capacity'):
                        ws.cell(row, 1, "  Capacity:")
                        ws.cell(row, 2, selected_filters.get('capacity'))
                        row += 1
                
                row += 1
                
                # Filtered cards count
                ws.cell(row, 1, "Filtered Product Cards Count:").font = Font(bold=True)
                ws.cell(row, 2, filtered_cards.get('card_count', 0))
                row += 1
                
                # Check if there's an error
                has_error = filtered_cards.get('error') is not None
                filtering_works = filtered_cards.get('filtering_works', False) and not has_error
                
                ws.cell(row, 1, "Filtering Works:")
                status_text = 'Fail' if has_error else ('Yes' if filtering_works else 'No')
                ws.cell(row, 2, status_text)
                ws.cell(row, 2).fill = PatternFill(start_color="C6EFCE" if filtering_works and not has_error else "FFC7CE", 
                                                   end_color="C6EFCE" if filtering_works and not has_error else "FFC7CE", 
                                                   fill_type="solid")
                row += 1
                
                # Show error details if there's an error
                if has_error:
                    ws.cell(row, 1, "Error Type:").font = Font(bold=True, color="FF0000")
                    ws.cell(row, 2, filtered_cards.get('error', 'Unknown Error'))
                    row += 1
                    if filtered_cards.get('error_message'):
                        ws.cell(row, 1, "Error Message:").font = Font(bold=True, color="FF0000")
                        ws.cell(row, 2, filtered_cards.get('error_message', ''))
                        row += 1
                
                row += 1
                
                # Filtered cards details table
                filtered_cards_list = filtered_cards.get('cards', [])
                if filtered_cards_list:
                    ws.cell(row, 1, "Filtered Product Cards Details").font = Font(bold=True, size=11)
                    row += 1
                    
                    headers = ['Index', 'Product Title', 'Title Font', 'Title Color', 'Description', 'Description Font', 
                              'Container Size', 'Image Src', 'Interface', 'Interface Font', 'Form Factor', 'Form Factor Font',
                              'Capacity', 'Capacity Font', 'View Details Link', 'View Details Button Text', 'Button Font', 
                              'Button BG Color', 'Compare Button Text', 'URL Format Valid']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row, col, header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    row += 1
                    
                    for card in filtered_cards_list:
                        ws.cell(row, 1, card.get('index', ''))
                        
                        # Title
                        title_text = card.get('title', {}).get('text', '')
                        ws.cell(row, 2, title_text)
                        ws.cell(row, 3, card.get('title', {}).get('fontSize', ''))
                        ws.cell(row, 4, card.get('title', {}).get('color', ''))
                        
                        # Description
                        desc_text = card.get('description', {}).get('text', '')
                        ws.cell(row, 5, desc_text[:100] + '...' if len(desc_text) > 100 else desc_text)
                        ws.cell(row, 6, self._format_font_size(card.get('description', {}).get('fontSize', '')))
                        
                        # Container
                        container = card.get('container', {})
                        container_width = float(container.get('width', 0))
                        container_height = float(container.get('height', 0))
                        container_size = f"{container_width:.2f}x{container_height:.2f}"
                        ws.cell(row, 7, container_size)
                        
                        # Image
                        image_src = card.get('image', {}).get('src', '')
                        ws.cell(row, 8, image_src[:50] + '...' if len(image_src) > 50 else image_src)
                        
                        # Interface
                        interface_text = card.get('interface', {}).get('text', '')
                        ws.cell(row, 9, interface_text)
                        ws.cell(row, 10, self._format_font_size(card.get('interface', {}).get('fontSize', '')))
                        
                        # Form Factor
                        form_factor_text = card.get('form_factor', {}).get('text', '')
                        ws.cell(row, 11, form_factor_text)
                        ws.cell(row, 12, self._format_font_size(card.get('form_factor', {}).get('fontSize', '')))
                        
                        # Capacity
                        capacity_text = card.get('capacity', {}).get('text', '')
                        ws.cell(row, 13, capacity_text)
                        ws.cell(row, 14, self._format_font_size(card.get('capacity', {}).get('fontSize', '')))
                        
                        # View Details
                        view_details_link = card.get('view_details_link', '')
                        ws.cell(row, 15, view_details_link)
                        view_details_btn = card.get('view_details_button', {})
                        ws.cell(row, 16, view_details_btn.get('text', ''))
                        ws.cell(row, 17, self._format_font_size(view_details_btn.get('fontSize', '')))
                        ws.cell(row, 18, view_details_btn.get('backgroundColor', ''))
                        
                        # Compare Button
                        compare_btn = card.get('compare_button', {})
                        ws.cell(row, 19, compare_btn.get('text', ''))
                        
                        # URL Format Valid
                        url_valid = card.get('url_format_valid', False)
                        ws.cell(row, 20, 'Yes' if url_valid else 'No')
                        ws.cell(row, 20).fill = PatternFill(start_color="C6EFCE" if url_valid else "FFC7CE", 
                                                           end_color="C6EFCE" if url_valid else "FFC7CE", 
                                                           fill_type="solid")
                        row += 1
                    
                    row += 1
        
        # Related Articles Section
        articles = results.get('related_articles', {})
        if articles.get('found'):
            ws.cell(row, 1, "Related Articles").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row, 1, "Article Cards Count:")
            ws.cell(row, 2, articles.get('card_count', 0))
            row += 2
            
            # Article cards details
            if articles.get('cards'):
                headers = ['Index', 'Title', 'Category', 'Link', 'URL Format Valid', 'URL Matches Title']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                row += 1
                
                for article in articles.get('cards', []):
                    ws.cell(row, 1, article.get('index', ''))
                    title_text = article.get('title', {}).get('text', '')
                    ws.cell(row, 2, title_text)
                    ws.cell(row, 3, article.get('category', ''))
                    ws.cell(row, 4, article.get('link', ''))
                    url_valid = article.get('url_format_valid', False)
                    ws.cell(row, 5, 'Yes' if url_valid else 'No')
                    ws.cell(row, 5).fill = PatternFill(start_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       end_color="C6EFCE" if url_valid else "FFC7CE", 
                                                       fill_type="solid")
                    url_matches = article.get('url_matches_title', False)
                    ws.cell(row, 6, 'Yes' if url_matches else 'No')
                    ws.cell(row, 6).fill = PatternFill(start_color="C6EFCE" if url_matches else "FFC7CE", 
                                                       end_color="C6EFCE" if url_matches else "FFC7CE", 
                                                       fill_type="solid")
                    row += 1
        
        # Auto-adjust column widths
        column_widths = {
            1: 10,   # Index
            2: 20,   # Product Title
            3: 12,   # Title Font
            4: 15,   # Title Color
            5: 40,   # Description
            6: 12,   # Description Font
            7: 15,   # Container Size
            8: 40,   # Image Src
            9: 20,   # Interface
            10: 12,  # Interface Font
            11: 20,  # Form Factor
            12: 12,  # Form Factor Font
            13: 15,  # Capacity
            14: 12,  # Capacity Font
            15: 40,  # View Details Link
            16: 20,  # View Details Button Text
            17: 12,  # Button Font
            18: 15,  # Button BG Color
            19: 20,  # Compare Button Text
            20: 15   # URL Format Valid
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

