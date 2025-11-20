"""
Excel Report Generator for Featured Products
"""
import os
from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class FeaturedProductsReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def generate_excel_report(self, results: Dict) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.output_dir}/featured_products_report_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Featured Products"

        # Header row
        headers = [
            "Card",
            "Title",
            "Description",
            "Link",
            "HTTP Status",
            "Valid",
            "Image Src",
            "Image WxH",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 2
        for card in results.get("cards", {}).get("cards", []):
            ws.cell(row=row, column=1, value=card.get("index", 0))
            ws.cell(row=row, column=2, value=(card.get("title", "") or "")[:70])
            ws.cell(row=row, column=3, value=(card.get("description", "") or "")[:100])

            link = card.get("link", {})
            ws.cell(row=row, column=4, value=link.get("href", ""))
            ws.cell(row=row, column=5, value=link.get("status_code", ""))
            ws.cell(row=row, column=6, value="Yes" if link.get("is_valid") else "No")

            img = card.get("image", {})
            ws.cell(row=row, column=7, value=img.get("src", ""))
            ws.cell(row=row, column=8, value=f"{img.get('width', 0)}x{img.get('height', 0)}")

            # Highlight invalid links
            if not link.get("is_valid", True):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

            row += 1

        # Set column widths
        widths = [8, 35, 60, 50, 14, 8, 50, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Summary sheet
        ws2 = wb.create_sheet("Summary", 1)
        ws2["A1"] = "FEATURED PRODUCTS"
        ws2["A1"].font = Font(bold=True, size=16, color="366092")
        ws2.merge_cells('A1:C1')

        row = 3
        summary = results.get('summary', {})
        ws2.cell(row=row, column=1, value="Total Cards:").font = Font(bold=True)
        ws2.cell(row=row, column=2, value=summary.get('total_cards', 0))
        row += 1
        ws2.cell(row=row, column=1, value="Title Exists:").font = Font(bold=True)
        ws2.cell(row=row, column=2, value='Yes' if summary.get('title_exists') else 'No')
        row += 1
        ws2.cell(row=row, column=1, value="Chevrons Working:").font = Font(bold=True)
        ws2.cell(row=row, column=2, value='Yes' if summary.get('chevrons_working') else 'No')

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15

        # Font Styles sheet
        ws3 = wb.create_sheet("Font Styles", 2)
        ws3['A1'] = "Card"
        ws3['B1'] = "Element"
        ws3['C1'] = "Font Size"
        ws3['D1'] = "Font Color"
        ws3['E1'] = "Font Weight"
        ws3['F1'] = "Font Family"
        ws3['G1'] = "Text Transform"
        ws3['H1'] = "Decoration"
        for cell in ['A1','B1','C1','D1','E1','F1','G1','H1']:
            ws3[cell].font = Font(bold=True, color="FFFFFF")
            ws3[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        r = 2
        for card in results.get('cards', {}).get('cards', []):
            styles = card.get('font_styles', {})
            for key in ['title','description','link']:
                s = styles.get(key)
                if s:
                    ws3.cell(row=r, column=1, value=card.get('index', 0))
                    ws3.cell(row=r, column=2, value=key.upper())
                    ws3.cell(row=r, column=3, value=s.get('fontSize',''))
                    ws3.cell(row=r, column=4, value=s.get('color',''))
                    ws3.cell(row=r, column=5, value=s.get('fontWeight',''))
                    ws3.cell(row=r, column=6, value=s.get('fontFamily',''))
                    ws3.cell(row=r, column=7, value=s.get('textTransform',''))
                    ws3.cell(row=r, column=8, value=s.get('textDecorationLine',''))
                    r += 1

        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 14
        ws3.column_dimensions['C'].width = 14
        ws3.column_dimensions['D'].width = 24
        ws3.column_dimensions['E'].width = 14
        ws3.column_dimensions['F'].width = 28
        ws3.column_dimensions['G'].width = 18
        ws3.column_dimensions['H'].width = 16

        wb.save(filename)
        print(f"\n[EXCEL] Featured Products report saved: {filename}")
        return filename


