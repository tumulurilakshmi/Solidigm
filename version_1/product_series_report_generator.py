"""
Excel Report Generator for Product Series Validation
"""
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ProductSeriesReportGenerator:
    def __init__(self):
        self.reports_dir = Path('reports')
        self.reports_dir.mkdir(exist_ok=True)
    
    def generate_excel_report(self, results: list) -> str:
        """Generate Excel report for product series validation"""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"product_series_validation_report_{timestamp}.xlsx"
        filepath = self.reports_dir / filename
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        self._create_summary_sheet(wb, results)
        
        # Create sheet for each series
        for result in results:
            series = result.get('series', 'Unknown')
            self._create_series_sheet(wb, result, series)
        
        wb.save(filepath)
        return str(filepath)
    
    def _create_summary_sheet(self, wb: Workbook, results: list):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header
        ws['A1'] = "Product Series Validation Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Generated:"
        ws['B3'] = time.strftime("%Y-%m-%d %H:%M:%S")
        
        # Table headers
        headers = ['Series', 'URL', 'Page Loaded', 'Title Found', 'Expected Products', 
                  'Found Products', 'All Found', 'Filters Working', 'Links Valid', 'Comparison Working']
        
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        row = 6
        for result in results:
            summary = result.get('summary', {})
            series = result.get('series', 'Unknown')
            url = result.get('url', '')
            
            ws.cell(row, 1, series)
            ws.cell(row, 2, url)
            ws.cell(row, 3, 'Yes' if summary.get('page_loaded') else 'No')
            ws.cell(row, 4, 'Yes' if summary.get('title_found') else 'No')
            ws.cell(row, 5, summary.get('expected_products', 0))
            ws.cell(row, 6, summary.get('found_products', 0))
            ws.cell(row, 7, 'Yes' if summary.get('all_products_found') else 'No')
            ws.cell(row, 8, 'Yes' if summary.get('filters_working') else 'No')
            ws.cell(row, 9, 'Yes' if summary.get('links_valid') else 'No')
            ws.cell(row, 10, 'Yes' if summary.get('comparison_working') else 'No')
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['B'].width = 50  # URL column wider
    
    def _create_series_sheet(self, wb: Workbook, result: dict, series: str):
        """Create detailed sheet for a series"""
        ws = wb.create_sheet(f"{series} Series")
        
        # Header
        ws['A1'] = f"{series} Series Validation Details"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Hero Component
        row = 3
        ws.cell(row, 1, "Hero Component").font = Font(bold=True, size=12)
        row += 1
        
        hero = result.get('hero', {})
        ws.cell(row, 1, "Hero Found:")
        ws.cell(row, 2, 'Yes' if hero.get('found') else 'No')
        row += 1
        
        # Container Size
        container = hero.get('container', {})
        if container.get('found'):
            ws.cell(row, 1, "Container Size:")
            ws.cell(row, 2, f"{container.get('width', 0)}x{container.get('height', 0)} px")
            row += 1
        
        # Background Image
        background = hero.get('background', {})
        if background.get('found'):
            ws.cell(row, 1, "Background Image:")
            ws.cell(row, 2, 'Yes' if background.get('has_desktop') else 'No')
            row += 1
            if background.get('desktop_image', {}).get('src'):
                ws.cell(row, 1, "Desktop Image URL:")
                ws.cell(row, 2, background['desktop_image']['src'])
                row += 1
                ws.cell(row, 1, "Desktop Image Size:")
                ws.cell(row, 2, f"{background['desktop_image'].get('width', 0)}x{background['desktop_image'].get('height', 0)}")
                row += 1
        
        # Breadcrumbs
        breadcrumbs = hero.get('breadcrumbs', {})
        if breadcrumbs.get('found'):
            ws.cell(row, 1, "Breadcrumbs Found:")
            ws.cell(row, 2, 'Yes')
            row += 1
            ws.cell(row, 1, "Breadcrumb Levels:")
            ws.cell(row, 2, breadcrumbs.get('total_levels', 0))
            row += 1
            ws.cell(row, 1, "All Clickable (except last):")
            ws.cell(row, 2, 'Yes' if breadcrumbs.get('all_clickable_except_last') else 'No')
            row += 1
            
            # Breadcrumb details table
            row += 1
            headers = ['Level', 'Text', 'Clickable', 'Is Last', 'Font Size', 'Font Color', 'Href']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            for level in breadcrumbs.get('levels', []):
                ws.cell(row, 1, level.get('level', ''))
                ws.cell(row, 2, level.get('text', ''))
                ws.cell(row, 3, 'Yes' if level.get('is_clickable') else 'No')
                ws.cell(row, 4, 'Yes' if level.get('is_last') else 'No')
                ws.cell(row, 5, level.get('font_size', ''))
                ws.cell(row, 6, level.get('font_color', ''))
                ws.cell(row, 7, level.get('href', ''))
                row += 1
        
        # Title
        title = hero.get('title', {})
        if title.get('found'):
            row += 1
            ws.cell(row, 1, "Hero Title").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row, 1, "Title Text:")
            ws.cell(row, 2, title.get('text', ''))
            row += 1
            ws.cell(row, 1, "Font Size:")
            ws.cell(row, 2, title.get('font_size', ''))
            row += 1
            ws.cell(row, 1, "Font Color:")
            ws.cell(row, 2, title.get('font_color', ''))
            row += 1
            ws.cell(row, 1, "Font Family:")
            ws.cell(row, 2, title.get('font_family', ''))
            row += 1
            ws.cell(row, 1, "Font Weight:")
            ws.cell(row, 2, title.get('font_weight', ''))
            row += 1
        
        # Description
        description = hero.get('description', {})
        if description.get('found'):
            row += 1
            ws.cell(row, 1, "Hero Description").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row, 1, "Description Text:")
            ws.cell(row, 2, description.get('text', '')[:500])  # Limit to 500 chars
            row += 1
            ws.cell(row, 1, "Font Size:")
            ws.cell(row, 2, description.get('font_size', ''))
            row += 1
            ws.cell(row, 1, "Font Color:")
            ws.cell(row, 2, description.get('font_color', ''))
            row += 1
            ws.cell(row, 1, "Font Family:")
            ws.cell(row, 2, description.get('font_family', ''))
            row += 1
            ws.cell(row, 1, "Font Weight:")
            ws.cell(row, 2, description.get('font_weight', ''))
            row += 2
        
        # Page Structure
        ws.cell(row, 1, "Page Structure").font = Font(bold=True, size=12)
        row += 1
        
        page_struct = result.get('page_structure', {})
        ws.cell(row, 1, "Page Loaded:")
        ws.cell(row, 2, 'Yes' if page_struct.get('page_loaded') else 'No')
        row += 1
        
        ws.cell(row, 1, "Title:")
        ws.cell(row, 2, page_struct.get('title_text', ''))
        row += 1
        
        ws.cell(row, 1, "Breadcrumbs:")
        ws.cell(row, 2, ' > '.join(page_struct.get('breadcrumbs', [])))
        row += 2
        
        # Products
        ws.cell(row, 1, "Products").font = Font(bold=True, size=12)
        row += 1
        
        products = result.get('products', {})
        ws.cell(row, 1, "Total Products Found:")
        ws.cell(row, 2, products.get('product_count', 0))
        row += 1
        
        ws.cell(row, 1, "Expected Products:")
        ws.cell(row, 2, ', '.join(products.get('expected_products', [])))
        row += 1
        
        ws.cell(row, 1, "Found Product IDs:")
        ws.cell(row, 2, ', '.join(products.get('found_product_ids', [])))
        row += 2
        
        # Product Details Table
        product_list = products.get('products', [])
        if product_list:
            headers = ['#', 'Product ID', 'Name', 'Interface', 'Form Factor', 'Capacity', 'View Details Link']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            row += 1
            for product in product_list:
                ws.cell(row, 1, product.get('index', ''))
                ws.cell(row, 2, product.get('id', ''))
                ws.cell(row, 3, product.get('name', ''))
                ws.cell(row, 4, product.get('interface', ''))
                ws.cell(row, 5, product.get('form_factor', ''))
                ws.cell(row, 6, product.get('capacity', ''))
                ws.cell(row, 7, product.get('view_details_link', ''))
                row += 1
        
        row += 1
        
        # Filters
        ws.cell(row, 1, "Filters").font = Font(bold=True, size=12)
        row += 1
        
        filters = result.get('filters', {})
        ws.cell(row, 1, "Filters Found:")
        ws.cell(row, 2, 'Yes' if filters.get('filters_found') else 'No')
        row += 1
        
        ws.cell(row, 1, "Interface Filter:")
        ws.cell(row, 2, 'Yes' if filters.get('interface_filter') else 'No')
        row += 1
        
        ws.cell(row, 1, "Form Factor Filter:")
        ws.cell(row, 2, 'Yes' if filters.get('form_factor_filter') else 'No')
        row += 1
        
        ws.cell(row, 1, "Capacity Filter:")
        ws.cell(row, 2, 'Yes' if filters.get('capacity_filter') else 'No')
        row += 2
        
        # Links
        ws.cell(row, 1, "Links").font = Font(bold=True, size=12)
        row += 1
        
        links = result.get('links', {})
        ws.cell(row, 1, "Total Links:")
        ws.cell(row, 2, links.get('total_links', 0))
        row += 1
        
        ws.cell(row, 1, "Valid Links:")
        ws.cell(row, 2, links.get('valid_links', 0))
        row += 1
        
        ws.cell(row, 1, "Invalid Links:")
        ws.cell(row, 2, links.get('invalid_links', 0))
        row += 2
        
        # Comparison
        ws.cell(row, 1, "Comparison Feature").font = Font(bold=True, size=12)
        row += 1
        
        comparison = result.get('comparison', {})
        ws.cell(row, 1, "Comparison Found:")
        ws.cell(row, 2, 'Yes' if comparison.get('comparison_found') else 'No')
        row += 1
        
        ws.cell(row, 1, "Max Products:")
        ws.cell(row, 2, comparison.get('max_products', 5))
        row += 2
        
        # Related Articles
        ws.cell(row, 1, "Related Articles").font = Font(bold=True, size=12)
        row += 1
        
        articles = result.get('articles', {})
        ws.cell(row, 1, "Section Found:")
        ws.cell(row, 2, 'Yes' if articles.get('section_found') else 'No')
        row += 1
        
        ws.cell(row, 1, "Article Count:")
        ws.cell(row, 2, articles.get('article_count', 0))
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50

