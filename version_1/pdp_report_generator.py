"""
Product Detail Page (PDP) Report Generator
Generates Excel report for PDP validation
"""
from typing import Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from base_report_generator import BaseReportGenerator
import os


class PDPReportGenerator(BaseReportGenerator):
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_excel_report(self, results: Dict) -> str:
        """Generate Excel report for PDP validation"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.output_dir}/pdp_report_{timestamp}.xlsx"
            
            print(f"\n[EXCEL] Generating report: {filename}")
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Summary Sheet
            try:
                self._create_summary_sheet(wb, results)
                print(f"[EXCEL] Summary sheet created")
            except Exception as e:
                print(f"[WARNING] Error creating summary sheet: {str(e)}")
            
            # Hero Sheet
            if results.get('hero'):
                try:
                    self._create_hero_sheet(wb, results['hero'])
                    print(f"[EXCEL] Hero sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating hero sheet: {str(e)}")
            
            # Cards Sheet
            if results.get('cards'):
                try:
                    self._create_cards_sheet(wb, results['cards'])
                    print(f"[EXCEL] Cards sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating cards sheet: {str(e)}")
            
            # Related Articles Sheet
            if results.get('related_articles'):
                try:
                    self._create_articles_sheet(wb, results['related_articles'])
                    print(f"[EXCEL] Related Articles sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating articles sheet: {str(e)}")
            
            # Search Sheet
            if results.get('search'):
                try:
                    self._create_search_sheet(wb, results['search'])
                    print(f"[EXCEL] Search sheet created")
                except Exception as e:
                    print(f"[WARNING] Error creating search sheet: {str(e)}")
            
            # Save the workbook
            try:
                wb.save(filename)
                print(f"\n[EXCEL] [OK] Report successfully saved: {filename}")
                return filename
            except Exception as e:
                print(f"\n[ERROR] Failed to save Excel file: {str(e)}")
                raise
        
        except Exception as e:
            print(f"\n[ERROR] Excel report generation failed: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = "PDP VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        ws.cell(row=row, column=1, value="Product URL:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('url', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="Product Name:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('product_name', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="Timestamp:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 2
        
        ws.cell(row=row, column=1, value="COMPONENT SUMMARY").font = Font(bold=True, size=12)
        row += 2
        
        ws.cell(row=row, column=1, value="Hero Component:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Found' if results.get('hero', {}).get('found') else 'Not Found')
        row += 1
        
        ws.cell(row=row, column=1, value="Header:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Found' if results.get('header_footer', {}).get('header_found') else 'Not Found')
        row += 1
        
        ws.cell(row=row, column=1, value="Footer:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Found' if results.get('header_footer', {}).get('footer_found') else 'Not Found')
        row += 1
        
        ws.cell(row=row, column=1, value="Cards Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('cards', {}).get('card_count', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Related Articles:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('related_articles', {}).get('article_count', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Search Component:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Found' if results.get('search', {}).get('component_exists') else 'Not Found')
        row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_hero_sheet(self, wb: Workbook, hero_data: Dict):
        """Create hero component sheet"""
        ws = wb.create_sheet("Hero")
        
        ws['A1'] = "HERO COMPONENT DETAILS"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Found:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if hero_data.get('found') else 'No')
        row += 1
        
        # Add more hero details as needed
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_cards_sheet(self, wb: Workbook, cards_data: Dict):
        """Create cards sheet"""
        ws = wb.create_sheet("Cards")
        
        ws['A1'] = "PRODUCT CARDS DETAILS"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:J1')
        
        row = 3
        ws.cell(row=row, column=1, value="Total Cards:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=cards_data.get('card_count', 0))
        row += 2
        
        # Cards table
        headers = ["Card #", "Title", "View Details Link", "Navigation Tested", "Navigation Success", "Compare Button"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for card in cards_data.get('cards', []):
            ws.cell(row=row, column=1, value=card.get('index', ''))
            ws.cell(row=row, column=2, value=card.get('title', ''))
            ws.cell(row=row, column=3, value=card.get('view_details_link', {}).get('href', ''))
            ws.cell(row=row, column=4, value='Yes' if card.get('navigation_tested') else 'No')
            ws.cell(row=row, column=5, value='Yes' if card.get('navigation_success') else 'No')
            ws.cell(row=row, column=6, value=card.get('compare_button', {}).get('text', ''))
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
    
    def _create_articles_sheet(self, wb: Workbook, articles_data: Dict):
        """Create related articles sheet"""
        ws = wb.create_sheet("Related Articles")
        
        ws['A1'] = "RELATED ARTICLES DETAILS"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        row = 3
        ws.cell(row=row, column=1, value="Article Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=articles_data.get('article_count', 0))
        row += 2
        
        # Articles table
        headers = ["Article #", "Title", "Link URL", "Image Source"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for article in articles_data.get('articles', []):
            ws.cell(row=row, column=1, value=article.get('index', ''))
            ws.cell(row=row, column=2, value=article.get('title', ''))
            ws.cell(row=row, column=3, value=article.get('link', {}).get('href', ''))
            ws.cell(row=row, column=4, value=article.get('image', {}).get('src', ''))
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
    
    def _create_search_sheet(self, wb: Workbook, search_data: Dict):
        """Create search component sheet"""
        ws = wb.create_sheet("Search")
        
        ws['A1'] = "SEARCH COMPONENT DETAILS"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        ws.cell(row=row, column=1, value="Component Found:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Yes' if search_data.get('component_exists') else 'No')
        row += 1
        
        if search_data.get('title', {}).get('text'):
            ws.cell(row=row, column=1, value="Title:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=search_data['title']['text'])
            row += 1
        
        ws.cell(row=row, column=1, value="Suggestions Count:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=search_data.get('suggestion_count', 0))
        row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

