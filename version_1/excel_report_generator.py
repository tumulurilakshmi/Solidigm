"""
Excel Report Generator for validation results
"""
import os
from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def generate_excel_report(self, validation_results: Dict) -> str:
        """Generate Excel report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.output_dir}/validation_report_{timestamp}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create Summary sheet
        self._create_summary_sheet(wb, validation_results)
        
        # Create UI Validation Details sheet
        self._create_ui_validation_sheet(wb, validation_results)
        
        # Create Link Validation sheet
        self._create_link_validation_sheet(wb, validation_results)
        
        # Create Failure Details sheet
        self._create_failure_details_sheet(wb, validation_results)
        
        wb.save(filename)
        print(f"[EXCEL] Report generated: {filename}")
        return filename
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        normal_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "SOLIDIGM VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:D1')
        
        # Validation Info
        row = 3
        if 'validation_info' in results:
            info = results['validation_info']
            ws['A3'] = "URL:"
            ws['B3'] = info.get('url', 'N/A')
            ws['A4'] = "Locale:"
            ws['B4'] = info.get('locale', 'N/A')
            ws['A5'] = "Timestamp:"
            ws['B5'] = info.get('timestamp', 'N/A')
            
            for cell in ['A3', 'B3', 'A4', 'B4', 'A5', 'B5']:
                ws[cell].font = normal_font
        
        # Check for error
        if 'error' in results:
            row = 7
            ws['A7'] = "ERROR OCCURRED"
            ws['A7'].font = Font(bold=True, size=12, color="DC3545")
            ws['A8'] = results['error']
            ws['A8'].font = normal_font
            ws.column_dimensions['A'].width = 80
        else:
            # Overall Summary
            summary = results.get('overall_summary', {})
            row = 8
            ws['A8'] = "OVERALL SUMMARY"
            ws['A8'].font = header_font
            ws['A8'].fill = header_fill
            ws.merge_cells('A8:B8')
            
            data = [
                ["Metric", "Value"],
                ["Total UI Validations", summary.get('total_ui_validations', 0)],
                ["UI Validations Passed", summary.get('passed_ui_validations', 0)],
                ["UI Validations Failed", summary.get('failed_ui_validations', 0)],
                ["Total Links", summary.get('total_links', 0)],
                ["Valid Links", summary.get('valid_links', 0)],
                ["Broken Links", summary.get('broken_links', 0)],
            ]
            
            for i, row_data in enumerate(data, start=9):
                for j, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = value
                    cell.font = normal_font
                    cell.border = border
                    
                    if i == 9:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    elif j == 2 and isinstance(value, int):
                        # Bold numbers
                        cell.font = Font(bold=True, size=11)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _create_ui_validation_sheet(self, wb: Workbook, results: Dict):
        """Create UI validation details sheet"""
        ws = wb.create_sheet("UI Validation Details", 1)
        
        if 'error' in results:
            ws['A1'] = "ERROR: " + results['error']
            return
        
        ui_result = results.get('ui_validation', {})
        summary = ui_result.get('summary', {})
        
        # Header
        ws['A1'] = "Category"
        ws['B1'] = "Total"
        ws['C1'] = "Passed"
        ws['D1'] = "Failed"
        ws['E1'] = "Pass %"
        ws['F1'] = "Status"
        
        # Apply header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = Alignment(horizontal='center')
        
        # Add data
        row = 2
        categories = summary.get('by_category', {})
        
        for category, data in categories.items():
            if data.get('total', 0) > 0:
                ws.cell(row=row, column=1, value=category.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=data['total'])
                ws.cell(row=row, column=3, value=data['passed'])
                ws.cell(row=row, column=4, value=data['failed'])
                
                # Calculate pass percentage
                pass_pct = round((data['passed'] / data['total'] * 100) if data['total'] > 0 else 0, 1)
                ws.cell(row=row, column=5, value=f"{pass_pct}%")
                
                # Status
                status = "PASS" if data['failed'] == 0 else "FAIL"
                ws.cell(row=row, column=6, value=status)
                
                # Apply colors
                if data['failed'] == 0:
                    fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                else:
                    fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.font = Font(size=10)
                
                row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def _create_link_validation_sheet(self, wb: Workbook, results: Dict):
        """Create link validation sheet"""
        ws = wb.create_sheet("Link Validation", 2)
        
        if 'error' in results:
            ws['A1'] = "ERROR: " + results['error']
            return
        
        link_result = results.get('link_validation', {})
        
        # Header
        ws['A1'] = "URL"
        ws['B1'] = "Text"
        ws['C1'] = "Status Code"
        ws['D1'] = "Is Valid"
        ws['E1'] = "Message"
        
        # Apply header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = Alignment(horizontal='center')
        
        # Add data
        row = 2
        broken_details = link_result.get('broken_details', [])
        
        for link in broken_details[:50]:  # Limit to 50 rows
            ws.cell(row=row, column=1, value=link.get('url', ''))
            ws.cell(row=row, column=2, value=link.get('text', ''))
            ws.cell(row=row, column=3, value=link.get('status_code', ''))
            ws.cell(row=row, column=4, value="NO" if not link.get('is_valid', True) else "YES")
            ws.cell(row=row, column=5, value=link.get('message', ''))
            
            # Apply color based on validity
            if not link.get('is_valid', True):
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            else:
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = Font(size=10)
            
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
    
    def _create_failure_details_sheet(self, wb: Workbook, results: Dict):
        """Create failure details sheet"""
        ws = wb.create_sheet("Failure Details", 3)
        
        if 'error' in results:
            ws['A1'] = "ERROR: " + results['error']
            return
        
        # Header
        ws['A1'] = "Category"
        ws['B1'] = "Failure Details"
        
        # Apply header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        
        # Add failure details
        row = 2
        ui_result = results.get('ui_validation', {})
        summary = ui_result.get('summary', {})
        
        for category, data in summary.get('by_category', {}).items():
            if data.get('failed', 0) > 0 and data.get('details', []):
                for detail in data['details'][:10]:  # Limit details
                    if isinstance(detail, dict) and 'details' in detail:
                        ws.cell(row=row, column=1, value=category.replace('_', ' ').title())
                        ws.cell(row=row, column=2, value=detail['details'])
                        
                        # Color for failures
                        fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        for col in [1, 2]:
                            cell = ws.cell(row=row, column=col)
                            cell.fill = fill
                            cell.font = Font(size=10)
                        
                        row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 80

