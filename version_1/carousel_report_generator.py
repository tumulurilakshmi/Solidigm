"""
Excel Report Generator for Carousel Validation
"""
import os
from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class CarouselReportGenerator:
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def generate_excel_report(self, results: Dict) -> str:
        """Generate Excel report for carousel validation"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.output_dir}/carousel_report_{timestamp}.xlsx"
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary Sheet
        self._create_summary_sheet(wb, results)
        
        # Carousel Details Sheet
        self._create_carousel_details_sheet(wb, results)
        
        # Font Styles Sheet
        self._create_font_styles_sheet(wb, results)
        
        # Progress Bar Sheet
        self._create_progress_bar_sheet(wb, results)
        
        wb.save(filename)
        print(f"\n[EXCEL] Report generated: {filename}")
        return filename
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = "CAROUSEL VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws.merge_cells('A1:B1')
        
        row = 3
        ws.cell(row=row, column=1, value="Total Carousels:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get('carousel_count', 0))
        row += 1
        
        # For each carousel
        for i, carousel in enumerate(results.get('carousels', []), 1):
            row += 1
            ws.cell(row=row, column=1, value=f"Carousel {i}:").font = Font(bold=True, size=11)
            row += 1
            ws.cell(row=row, column=2, value=f"Total Cards: {carousel.get('slide_count', 0)}")
            row += 1
            container = carousel.get('container', {})
            ws.cell(row=row, column=2, value=f"Container: {container.get('width', 0)}x{container.get('height', 0)}")
            
            # Add button count summary
            slides = carousel.get('slides', [])
            if slides:
                bg_images_present = sum(1 for s in slides if s.get('background_image_present'))
                row += 1
                ws.cell(row=row, column=2, value=f"Background Images: {bg_images_present}/{len(slides)} cards")
                
                button_summary = {}
                for slide in slides:
                    count = slide.get('button_count', 0)
                    button_summary[count] = button_summary.get(count, 0) + 1
                if button_summary:
                    summary_text = ', '.join([f"{count} button(s): {freq} card(s)" for count, freq in sorted(button_summary.items())])
                    row += 1
                    ws.cell(row=row, column=2, value=f"Button Distribution: {summary_text}")
            
            row += 1
            nav = carousel.get('navigation', {})
            nav_text = f"Navigation: Left={nav.get('left_chevron_visible')}, Right={nav.get('right_chevron_visible')}"
            if nav.get('left_clicks_tested', 0) > 0:
                nav_text += f" (Left clicks: {nav.get('left_clicks_successful', 0)}/{nav.get('left_clicks_tested', 0)})"
            if nav.get('right_clicks_tested', 0) > 0:
                nav_text += f" (Right clicks: {nav.get('right_clicks_successful', 0)}/{nav.get('right_clicks_tested', 0)})"
            ws.cell(row=row, column=2, value=nav_text)
            row += 1
            pb = carousel.get('progress_bar', {})
            ws.cell(row=row, column=2, value=f"Progress Bar: {'Yes' if pb.get('exists') else 'No'} ({pb.get('indicator_count', 0)} indicators)")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def _create_carousel_details_sheet(self, wb: Workbook, results: Dict):
        """Create carousel details sheet"""
        ws = wb.create_sheet("Carousel Details", 1)
        
        ws['A1'] = "Carousel"
        ws['B1'] = "Slide"
        ws['C1'] = "Title"
        ws['D1'] = "Description"
        ws['E1'] = "Buttons (text, link, status)"
        ws['F1'] = "Button Count"
        ws['G1'] = "BG Image"
        ws['H1'] = "BG Image Present"
        ws['I1'] = "Link Status"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for i, carousel in enumerate(results.get('carousels', []), 1):
            for slide in carousel.get('slides', []):
                ws.cell(row=row, column=1, value=f"Carousel {i}")
                ws.cell(row=row, column=2, value=slide.get('index', 0))
                ws.cell(row=row, column=3, value=slide.get('title', '')[:50])
                ws.cell(row=row, column=4, value=slide.get('description', '')[:80])
                
                # Button information
                buttons = slide.get('buttons', [])
                button_count = slide.get('button_count', len(buttons))
                
                # Create detailed button info (include link and status code)
                button_details = []
                for i, btn in enumerate(buttons[:2]):  # Max 2 buttons
                    btn_text = btn.get('text', '')
                    # Prefer relative link if available
                    btn_href = btn.get('href', '') or btn.get('href_absolute', '')
                    status_code = btn.get('status_code')
                    # Prefer HTTP status validity; fallback to click-based navigation
                    if btn.get('is_valid') is True:
                        status = 'OK'
                    elif btn.get('is_valid') is False:
                        status = 'BAD'
                    else:
                        status = 'OK' if btn.get('navigates_correctly') else 'BAD' if btn.get('click_tested') else '?'
                    status_suffix = f" {status}{'' if status_code is None else f' [{status_code}]'}"
                    button_details.append(f"{btn_text} ({btn_href}){status_suffix}")
                
                btn_text = ' | '.join(button_details)[:80]
                
                ws.cell(row=row, column=5, value=btn_text)
                ws.cell(row=row, column=6, value=button_count)
                
                # Background image
                bg_images = slide.get('background_images', [])
                bg_image = slide.get('background_image', '')
                bg_present = slide.get('background_image_present', False)
                
                # Show all background images if multiple
                if len(bg_images) > 1:
                    bg_display = f"{bg_image[:40]} (+{len(bg_images)-1} more)"
                else:
                    bg_display = bg_image[:60]
                
                ws.cell(row=row, column=7, value=bg_display)
                ws.cell(row=row, column=8, value='Yes' if bg_present else 'No')
                
                # Link validation status
                links = slide.get('links', [])
                if links:
                    valid_links = sum(1 for l in links if l.get('is_valid', False))
                    link_status = f"{valid_links}/{len(links)} valid"
                else:
                    link_status = "No links"
                ws.cell(row=row, column=9, value=link_status)
                
                # Color coding for issues
                if not bg_present:
                    # Yellow for missing background image
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif any(not btn.get('navigates_correctly', True) for btn in buttons if btn.get('click_tested')):
                    # Red for navigation issues
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                
                row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 60
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
    
    def _create_font_styles_sheet(self, wb: Workbook, results: Dict):
        """Create font styles sheet"""
        ws = wb.create_sheet("Font Styles", 2)
        
        ws['A1'] = "Carousel"
        ws['B1'] = "Element"
        ws['C1'] = "Font Size"
        ws['D1'] = "Font Color"
        
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for i, carousel in enumerate(results.get('carousels', []), 1):
            for fs in carousel.get('font_styles', []):
                ws.cell(row=row, column=1, value=f"Carousel {i}")
                ws.cell(row=row, column=2, value=fs.get('element', '').upper())
                ws.cell(row=row, column=3, value=fs.get('font_size', ''))
                ws.cell(row=row, column=4, value=fs.get('font_color', ''))
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                
                row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
    
    def _create_progress_bar_sheet(self, wb: Workbook, results: Dict):
        """Create progress bar sheet"""
        ws = wb.create_sheet("Progress Bar", 3)
        
        ws['A1'] = "Carousel"
        ws['B1'] = "Exists"
        ws['C1'] = "Visible"
        ws['D1'] = "Indicators"
        ws['E1'] = "Active Slide Index"
        ws['F1'] = "Active Pagination Index"
        ws['G1'] = "Active Matches"
        
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for i, carousel in enumerate(results.get('carousels', []), 1):
            ws.cell(row=row, column=1, value=f"Carousel {i}")
            
            pb = carousel.get('progress_bar', {})
            ws.cell(row=row, column=2, value='Yes' if pb.get('exists') else 'No')
            ws.cell(row=row, column=3, value='Yes' if pb.get('is_visible') else 'No')
            ws.cell(row=row, column=4, value=pb.get('indicator_count', 0))
            ws.cell(row=row, column=5, value=pb.get('active_slide_index', ''))
            ws.cell(row=row, column=6, value=pb.get('active_pagination_index', ''))
            ws.cell(row=row, column=7, value='Yes' if pb.get('active_matches_slide') else 'No')
            
            if pb.get('exists'):
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 24
        ws.column_dimensions['G'].width = 16
