"""
Comprehensive Validation Script
Validates Homepage, Data Center page, and all Product Series pages (D3, D5, D7) at once
"""
import sys
import json
from pathlib import Path
from playwright.sync_api import sync_playwright
from homepage_validator import HomePageValidator
from home_page_report_generator import HomePageReportGenerator
from data_center_page_validator import DataCenterPageValidator
from data_center_page_report_generator import DataCenterPageReportGenerator
from product_series_validator import ProductSeriesValidator
from product_series_report_generator import ProductSeriesReportGenerator
from datetime import datetime


def main():
    print("=" * 100)
    print(" " * 20 + "COMPREHENSIVE VALIDATION - ALL PAGES")
    print("=" * 100)
    print("\nThis will validate:")
    print("  1. Homepage")
    print("  2. Data Center page")
    print("  3. D3 Series page")
    print("  4. D5 Series page")
    print("  5. D7 Series page")
    print("\n" + "=" * 100)
    
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=False, args=['--no-sandbox'])
    page = browser.new_page(viewport={'width': 1920, 'height': 1080})
    page.set_default_timeout(120000)
    
    all_results = {
        'homepage': {},
        'data_center': {},
        'd3': {},
        'd5': {},
        'd7': {},
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    try:
        # 1. Validate Homepage
        print("\n" + "="*100)
        print(" " * 40 + "1. HOMEPAGE VALIDATION")
        print("="*100)
        try:
            homepage_url = "https://www.solidigm.com"
            homepage_validator = HomePageValidator(page, homepage_url)
            homepage_results = homepage_validator.validate_complete_homepage()
            all_results['homepage'] = homepage_results
            
            # Generate homepage report
            if 'error' not in homepage_results:
                try:
                    homepage_report_gen = HomePageReportGenerator()
                    homepage_report = homepage_report_gen.generate_excel_report(homepage_results)
                    print(f"\n[SUCCESS] Homepage report saved: {homepage_report}")
                    all_results['homepage']['report_file'] = homepage_report
                except Exception as e:
                    print(f"\n[ERROR] Homepage report generation failed: {str(e)}")
        except Exception as e:
            print(f"\n[ERROR] Homepage validation failed: {str(e)}")
            all_results['homepage']['error'] = str(e)
        
        page.wait_for_timeout(2000)
        
        # 2. Validate Data Center page
        print("\n" + "="*100)
        print(" " * 35 + "2. DATA CENTER PAGE VALIDATION")
        print("="*100)
        try:
            data_center_url = "https://www.solidigm.com/products/data-center.html"
            data_center_validator = DataCenterPageValidator(page)
            data_center_results = data_center_validator.validate_data_center_page(data_center_url)
            all_results['data_center'] = data_center_results
            
            # Generate data center report
            if 'error' not in data_center_results:
                try:
                    dc_report_gen = DataCenterPageReportGenerator()
                    dc_report = dc_report_gen.generate_excel_report(data_center_results)
                    print(f"\n[SUCCESS] Data Center report saved: {dc_report}")
                    all_results['data_center']['report_file'] = dc_report
                except Exception as e:
                    print(f"\n[ERROR] Data Center report generation failed: {str(e)}")
        except Exception as e:
            print(f"\n[ERROR] Data Center validation failed: {str(e)}")
            all_results['data_center']['error'] = str(e)
        
        page.wait_for_timeout(2000)
        
        # 3-5. Validate Product Series pages (D3, D5, D7)
        series_to_validate = ['D3', 'D5', 'D7']
        series_data_path = 'product_series.json'
        
        # Load series data
        try:
            with open(series_data_path, 'r', encoding='utf-8') as f:
                series_data = json.load(f)
        except Exception as e:
            print(f"[WARNING] Could not load series data: {str(e)}")
            series_data = {'product_series': []}
        
        for series_name in series_to_validate:
            print("\n" + "="*100)
            print(f" " * 35 + f"3-5. {series_name} SERIES VALIDATION")
            print("="*100)
            
            try:
                # Find series info
                series_info = None
                for s in series_data.get('product_series', []):
                    if s.get('series') == series_name:
                        series_info = s
                        break
                
                if not series_info:
                    print(f"[ERROR] Series '{series_name}' not found in series data")
                    all_results[series_name.lower()]['error'] = f"Series data not found"
                    continue
                
                series_url = series_info['url']
                series_validator = ProductSeriesValidator(page, series_data_path)
                series_results = series_validator.validate_series_page(series_url, series_name)
                series_results['series_info'] = series_info
                all_results[series_name.lower()] = series_results
                
                # Generate series report (individual)
                if 'error' not in series_results:
                    try:
                        series_report_gen = ProductSeriesReportGenerator()
                        series_report = series_report_gen.generate_excel_report([series_results])
                        print(f"\n[SUCCESS] {series_name} Series report saved: {series_report}")
                        all_results[series_name.lower()]['report_file'] = series_report
                    except Exception as e:
                        print(f"\n[ERROR] {series_name} Series report generation failed: {str(e)}")
                
                page.wait_for_timeout(2000)
                
            except Exception as e:
                print(f"\n[ERROR] {series_name} Series validation failed: {str(e)}")
                all_results[series_name.lower()]['error'] = str(e)
        
        # Generate combined summary report
        print("\n" + "="*100)
        print(" " * 35 + "GENERATING COMBINED SUMMARY")
        print("="*100)
        try:
            _generate_combined_summary(all_results)
        except Exception as e:
            print(f"\n[WARNING] Combined summary generation failed: {str(e)}")
        
        # Print overall summary
        print("\n" + "="*100)
        print("VALIDATION COMPLETE - SUMMARY")
        print("="*100)
        print(f"\nHomepage: {'✓ Validated' if all_results['homepage'].get('summary') else '✗ Failed'}")
        print(f"Data Center: {'✓ Validated' if all_results['data_center'].get('hero') else '✗ Failed'}")
        print(f"D3 Series: {'✓ Validated' if all_results['d3'].get('series') else '✗ Failed'}")
        print(f"D5 Series: {'✓ Validated' if all_results['d5'].get('series') else '✗ Failed'}")
        print(f"D7 Series: {'✓ Validated' if all_results['d7'].get('series') else '✗ Failed'}")
        
        print("\n" + "="*100)
        print("All individual reports have been generated.")
        print("="*100)
        
    except Exception as e:
        print(f"\n[ERROR] Comprehensive validation failed: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        browser.close()


def _generate_combined_summary(all_results: dict):
    """Generate a combined summary report"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/comprehensive_validation_summary_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "COMPREHENSIVE VALIDATION SUMMARY"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:B1')
    
    row = 3
    ws.cell(row=row, column=1, value="Validation Timestamp:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=all_results.get('timestamp', ''))
    row += 2
    
    # Homepage Summary
    ws.cell(row=row, column=1, value="HOMEPAGE").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    homepage = all_results.get('homepage', {})
    ws.cell(row=row, column=1, value="Status:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Validated' if homepage.get('summary') else 'Failed')
    row += 1
    
    if homepage.get('report_file'):
        ws.cell(row=row, column=1, value="Report File:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=homepage['report_file'])
        row += 1
    row += 1
    
    # Data Center Summary
    ws.cell(row=row, column=1, value="DATA CENTER PAGE").font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    dc = all_results.get('data_center', {})
    ws.cell(row=row, column=1, value="Status:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Validated' if dc.get('hero') else 'Failed')
    row += 1
    
    if dc.get('report_file'):
        ws.cell(row=row, column=1, value="Report File:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=dc['report_file'])
        row += 1
    row += 1
    
    # Series Pages Summary
    for series in ['D3', 'D5', 'D7']:
        ws.cell(row=row, column=1, value=f"{series} SERIES").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        series_data = all_results.get(series.lower(), {})
        ws.cell(row=row, column=1, value="Status:").font = Font(bold=True)
        ws.cell(row=row, column=2, value='Validated' if series_data.get('series') else 'Failed')
        row += 1
        
        if series_data.get('report_file'):
            ws.cell(row=row, column=1, value="Report File:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=series_data['report_file'])
            row += 1
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 70
    
    wb.save(filename)
    print(f"\n[SUCCESS] Combined summary report saved: {filename}")


if __name__ == "__main__":
    main()

