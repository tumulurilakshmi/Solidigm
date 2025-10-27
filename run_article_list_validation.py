"""
Run Article List validation and generate Excel report
"""
from playwright.sync_api import sync_playwright
from article_list_validator import ArticleListValidator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os


def generate_excel_report(results: dict):
    """Generate Excel report for article list"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/article_list_report_{timestamp}.xlsx"
    
    if not os.path.exists("reports"):
        os.makedirs("reports")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Summary Sheet
    ws = wb.create_sheet("Summary", 0)
    ws['A1'] = "ARTICLE LIST VALIDATION REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    row = 3
    summary = results.get('summary', {})
    ws.cell(row=row, column=1, value="Total Cards:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=summary.get('total_cards', 0))
    row += 1
    
    ws.cell(row=row, column=1, value="Title Exists:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Yes' if summary.get('title_exists') else 'No')
    row += 1
    
    ws.cell(row=row, column=1, value="View All Link Valid:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Yes' if summary.get('view_all_link_valid') else 'No')
    row += 1
    
    ws.cell(row=row, column=1, value="Chevrons Working:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Yes' if summary.get('chevrons_working') else 'No')
    row += 1
    
    ws.cell(row=row, column=1, value="Hover Working:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Yes' if summary.get('hover_working') else 'No')
    row += 1
    
    ws.cell(row=row, column=1, value="All Links Valid:").font = Font(bold=True)
    ws.cell(row=row, column=2, value='Yes' if summary.get('all_links_valid') else 'No')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    
    # Article Cards Sheet
    ws = wb.create_sheet("Article Cards")
    ws['A1'] = "Card"
    ws['B1'] = "Title"
    ws['C1'] = "Category"
    ws['D1'] = "Link"
    ws['E1'] = "Container Size"
    ws['F1'] = "Image Size"
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    row = 2
    for card in results.get('cards', {}).get('cards', []):
        ws.cell(row=row, column=1, value=card.get('index', 0))
        ws.cell(row=row, column=2, value=card.get('title', ''))
        ws.cell(row=row, column=3, value=card.get('category', ''))
        ws.cell(row=row, column=4, value=card.get('link', ''))
        
        container = card.get('container', {})
        ws.cell(row=row, column=5, value=f"{container.get('width', 0)}x{container.get('height', 0)}")
        
        img = card.get('image', {})
        ws.cell(row=row, column=6, value=f"{img.get('width', 0)}x{img.get('height', 0)}")
        
        row += 1
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    
    # Font Styles Sheet
    ws = wb.create_sheet("Font Styles")
    ws['A1'] = "Card"
    ws['B1'] = "Element"
    ws['C1'] = "Font Size"
    ws['D1'] = "Font Color"
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    row = 2
    for card in results.get('cards', {}).get('cards', []):
        for element_type, styles in card.get('font_styles', {}).items():
            ws.cell(row=row, column=1, value=card.get('index', 0))
            ws.cell(row=row, column=2, value=element_type.upper())
            ws.cell(row=row, column=3, value=styles.get('fontSize', ''))
            ws.cell(row=row, column=4, value=styles.get('color', ''))
            row += 1
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    wb.save(filename)
    print(f"\n[EXCEL] Report saved: {filename}")
    return filename


def main():
    url = "https://www.solidigm.com/"
    
    print("=" * 80)
    print(" " * 20 + "ARTICLE LIST VALIDATION")
    print("=" * 80)
    
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=False, args=['--no-sandbox'])
    page = browser.new_page(viewport={'width': 1920, 'height': 1080})
    page.set_default_timeout(90000)
    
    try:
        print(f"\n[INFO] Navigating to {url}...")
        page.goto(url, timeout=90000, wait_until='domcontentloaded')
        page.wait_for_timeout(5000)  # Wait for dynamic content
        
        title = page.title()
        print(f"[OK] Page loaded: {title}")
        
        # Run article list validation
        validator = ArticleListValidator(page)
        results = validator.validate_article_list()
        
        # Generate Excel report if validation succeeded
        if 'error' not in results and results.get('found'):
            generate_excel_report(results)
        
        print("\n" + "="*80)
        print("VALIDATION COMPLETE")
        print("="*80)
        
    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
    finally:
        browser.close()


if __name__ == "__main__":
    main()

