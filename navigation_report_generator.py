"""
Navigation (Header) Report Generator
"""
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from base_report_generator import BaseReportGenerator


class NavigationReportGenerator(BaseReportGenerator):
    """Generate Excel report for Navigation/Header component"""
    
    def create_sheet(self, wb: Workbook, nav_results: Dict):
        """Create detailed navigation sheet"""
        ws = wb.create_sheet("Navigation")
        
        # Summary Section
        ws['A1'] = "NAVIGATION MENU SUMMARY"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:D1')
        
        summary = nav_results.get('summary', {})
        row = 3
        
        ws.cell(row=row, column=1, value="Total Main Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_main_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Visible Main Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('visible_main_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Sub-Menu Items:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_sub_menu_items', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Total Links Checked:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary.get('total_links_checked', 0))
        row += 1
        
        ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
        valid_links = summary.get('valid_links', 0)
        ws.cell(row=row, column=2, value=valid_links)
        if valid_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        row += 1
        
        ws.cell(row=row, column=1, value="Broken Links:").font = Font(bold=True)
        broken_links = summary.get('broken_links', 0)
        ws.cell(row=row, column=2, value=broken_links)
        if broken_links > 0:
            ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 2
        
        # Main Menu Details Table
        ws.cell(row=row, column=1, value="MAIN MENU DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ["Menu Name", "Display Text", "Is Visible", "Has Mega Menu", "Sub-Menu Count", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        # Get sub-menu counts
        sub_menus = nav_results.get('sub_menus', {})
        main_menu_items = nav_results.get('main_menu_items', [])
        
        for menu in main_menu_items:
            menu_name = menu.get('name', '')
            sub_count = len(sub_menus.get(menu_name, []))
            
            ws.cell(row=row, column=1, value=menu_name)
            ws.cell(row=row, column=2, value=menu.get('text', ''))
            ws.cell(row=row, column=3, value='Yes' if menu.get('is_visible', False) else 'No')
            ws.cell(row=row, column=4, value='Yes' if menu.get('has_mega_menu', False) else 'No')
            ws.cell(row=row, column=5, value=sub_count)
            ws.cell(row=row, column=6, value=menu.get('status', ''))
            
            # Color coding
            if menu.get('status') == 'PASS':
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            else:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
            
            row += 1
        
        row += 1
        
        # Sub-Menu Details Table
        ws.cell(row=row, column=1, value="SUB-MENU DETAILS").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ["Main Menu", "Link Text", "URL", "Status Code", "Is Visible", "Link Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        row += 1
        
        for menu_name, sub_items in sub_menus.items():
            for item in sub_items:
                ws.cell(row=row, column=1, value=menu_name)
                ws.cell(row=row, column=2, value=item.get('text', ''))
                ws.cell(row=row, column=3, value=item.get('href', ''))
                
                status_code = item.get('status_code', 0)
                status_cell = ws.cell(row=row, column=4, value=status_code)
                status_cell.number_format = '0'
                
                ws.cell(row=row, column=5, value='Yes' if item.get('is_visible', False) else 'No')
                
                link_status = 'Working' if status_code == 200 else 'Broken' if status_code > 0 else 'Not Checked'
                ws.cell(row=row, column=6, value=link_status)
                
                # Color coding
                if status_code == 200:
                    fill_color = "D4EDDA"  # Green
                elif status_code == 0:
                    fill_color = "FFF3CD"  # Yellow
                else:
                    fill_color = "F8D7DA"  # Red
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        row += 1
        
        # Navigation Links Summary
        link_validation = nav_results.get('link_validation', {})
        if link_validation:
            ws.cell(row=row, column=1, value="NAVIGATION LINKS SUMMARY").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            ws.cell(row=row, column=1, value="Total Links Checked:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=link_validation.get('total_checked', 0))
            row += 1
            
            ws.cell(row=row, column=1, value="Valid Links:").font = Font(bold=True)
            valid = link_validation.get('valid_links', 0)
            ws.cell(row=row, column=2, value=valid)
            if valid > 0:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            row += 1
            
            ws.cell(row=row, column=1, value="Broken Links (4xx/5xx):").font = Font(bold=True)
            broken = link_validation.get('broken_links', 0)
            ws.cell(row=row, column=2, value=broken)
            if broken > 0:
                ws.cell(row=row, column=2).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            row += 1
            
            # Not Checked Links (timeouts/errors)
            not_checked = link_validation.get('not_checked_links', 0)
            if not_checked > 0:
                ws.cell(row=row, column=1, value="Not Checked (Timeouts/Errors):").font = Font(bold=True)
                ws.cell(row=row, column=2, value=not_checked)
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                row += 1
            
            row += 1
            
            # Broken Links Details Table (only actual broken links)
            broken_details = link_validation.get('broken_details', [])
            if broken_details:
                ws.cell(row=row, column=1, value="BROKEN LINKS DETAILS").font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
                
                headers = ["Link Text", "URL", "Status Code", "Is Visible", "Error Type"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                row += 1
                
                for link in broken_details:
                    ws.cell(row=row, column=1, value=link.get('text', '')[:50])
                    ws.cell(row=row, column=2, value=link.get('href', '')[:80])
                    status_code = link.get('status_code', 0)
                    ws.cell(row=row, column=3, value=status_code)
                    ws.cell(row=row, column=4, value='Yes' if link.get('is_visible', False) else 'No')
                    ws.cell(row=row, column=5, value='404 Not Found' if status_code == 404 else '403 Forbidden' if status_code == 403 else '500 Server Error' if status_code >= 500 else 'Other Error')
                    
                    # Color coding
                    if status_code == 404:
                        fill_color = "FFE4E1"  # Light red
                    elif status_code == 403:
                        fill_color = "FFB6C1"  # Pink
                    elif status_code >= 500:
                        fill_color = "F8D7DA"  # Red
                    else:
                        fill_color = "FFF3CD"  # Yellow
                    
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    row += 1
            row += 1
        
        # Font Styles Section
        font_styles = nav_results.get('font_styles', {})
        if font_styles.get('main_menu') or font_styles.get('sub_menu'):
            row += 1
            ws.cell(row=row, column=1, value="FONT STYLES").font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            headers = ["Element Type", "Name", "Font Size", "Font Color"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            row += 1
            
            # Main menu font styles
            for item in font_styles.get('main_menu', []):
                ws.cell(row=row, column=1, value="Main Menu")
                ws.cell(row=row, column=2, value=item.get('name', ''))
                ws.cell(row=row, column=3, value=self._format_font_size(item.get('font_size', '')))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                row += 1
            
            # Sub menu font styles
            for item in font_styles.get('sub_menu', []):
                ws.cell(row=row, column=1, value="Sub Menu")
                ws.cell(row=row, column=2, value=item.get('type', ''))
                ws.cell(row=row, column=3, value=self._format_font_size(item.get('font_size', '')))
                ws.cell(row=row, column=4, value=item.get('font_color', ''))
                row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

